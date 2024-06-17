import os
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO

def extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder, image_width=100, image_height=100):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # 「写真番号」と書かれているセルを探す
    photo_number_positions = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "写真番号":
                photo_number_positions.append((cell.row, cell.column))
                #print(f"写真番号 Row: {cell.row}, Column: {cell.column}")

    # 写真番号セルから右方向に移動した位置の値を取得するための辞書
    photo_name_dict = {}
    for p_row, p_col in photo_number_positions:
        for offset in range(1, 10):  # 最大10列までチェックする
            photo_name_cell = sheet.cell(row=p_row, column=p_col + offset)
            if photo_name_cell.value:
                photo_name_dict[(p_row, p_col)] = photo_name_cell.value
                #print(f"Photo name at ({p_row}, {p_col}): {photo_name_cell.value}")
                break

    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))

    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 写真番号セルから右方向の値で画像を保存
    used_photo_names = set()  # 既に使用された写真番号のセット
    
    for row, col, img in images:
        photo_name = None
        for (p_row, p_col), name in photo_name_dict.items():
            if row == p_row and col > p_col:
                photo_name = name
                break
        
        if not photo_name:
            # 指定された位置に写真番号がない場合、デフォルトの名前を使用
            photo_name = f"photo_{row}_{col}"
        
        # 重複しないファイル名を生成
        original_photo_name = photo_name
        count = 1
        while photo_name in used_photo_names:
            photo_name = f"{original_photo_name}_{count}"
            count += 1
        used_photo_names.add(photo_name)

        # 画像を保存する
        img_data = img._data()
        image_file_path = os.path.join(output_folder, f"{photo_name}.jpg")
        
        with open(image_file_path, "wb") as img_file:
            img_file.write(img_data)

        # 画像ファイルを読み込んでサイズを変更する
        img_object = Image(image_file_path)
        img_object.width = image_width
        img_object.height = image_height
        
        # 画像をExcelシートに追加する場合
        # img.anchor = f'{openpyxl.utils.get_column_letter(col)}{row}'
        # sheet.add_image(img_object)

    # 関数の最後でphoto_name_dictを返す
    return photo_name_dict

# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定
image_width = 1200  # 画像の幅を指定
image_height = 900  # 画像の高さを指定

# 関数の戻り値を取得
photo_name_dict = extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder, image_width, image_height)

# 保存先フルパスの表示
full_output_folder_path = os.path.abspath(output_folder)
print(f"Images are saved in: {full_output_folder_path}")

# フォルダパス
folder_path = full_output_folder_path

# 辞書の値をリストにして順番に取得
new_names = list(photo_name_dict.values())
print(new_names)  # [1, 2, '前回写真-3', ..., 23]

# フォルダ内のファイルをリストする
files = os.listdir(folder_path)
files = [f for f in files if f.startswith('photo_') and f.endswith('.jpg')]

# 古い名前に基づき写真ファイルを順番にソート
files.sort(key=lambda x: (int(x.split('_')[1]), int(x.split('_')[2].split('.')[0])))

# インデックスを初期化
index = 0

for file_name in files:
    # 'photo_' で始まり、'.jpg' で終わるファイルのみを対象にする
    if file_name.startswith('photo_') and file_name.endswith('.jpg'):
        full_path = os.path.join(folder_path, file_name)
        
        try:
            # 新しい名前を取得しインデックスを更新
            new_name = f"{new_names[index]}.jpg"
            new_full_path = os.path.join(folder_path, new_name)
            
            os.rename(full_path, new_full_path)
            print(f"Renamed {full_path} to {new_full_path}")
            
            # インデックスを次に進める
            index += 1

        except Exception as e:
            print(f"Error processing file {file_name}: {e}")

        # もしファイル数が辞書の名前数を超えるならばループを抜ける
        if index >= len(new_names):
            break
