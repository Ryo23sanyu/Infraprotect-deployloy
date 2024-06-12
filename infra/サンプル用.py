import os
import openpyxl
from openpyxl.drawing.image import Image

def extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))

    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 「写真番号」と書かれているセルを探す
    photo_number_positions = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "写真番号":
                photo_number_positions.append((cell.row, cell.column))

    # 写真番号セルから右方向に移動した位置の値を取得するための辞書作成
    photo_name_dict = {}
    for p_row, p_col in photo_number_positions:
        for offset in range(1, 10):  # 最大10列までチェックする
            photo_name_cell = sheet.cell(row=p_row, column=p_col + offset)
            if photo_name_cell.value:
                photo_name_dict[(p_row, p_col)] = photo_name_cell.value
                print(photo_name_dict)

                # 画像を保存する
                img_data = img._data()
                image_file_path = os.path.join(output_folder, f"画像：{photo_name_dict}.jpg")
                
                with open(image_file_path, "wb") as img_file:
                    img_file.write(img_data)

# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定

extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder)
