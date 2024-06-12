import os
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

def extract_images_from_sheet(excel_file, sheet_name, output_folder, resize_to=None):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]
    
    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))
    
    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 画像を保存するためのカウンター
    image_counter = 1
    
    for row, col, img in images:
        # 画像データをBytesIOに保存
        img_data = BytesIO(img._data())

        # Pillowで画像を開く
        pil_image = PILImage.open(img_data)
        
        # リサイズ処理（オプション）
        if resize_to:
            pil_image = pil_image.resize(resize_to, PILImage.Resampling.LANCZOS)
        
        # 写真番号を持つセルの行・列
        photo_number_row = row
        photo_number_col = col
        print(f"x:{photo_number_row}、Y:{photo_number_col}")
        
        # 列を1つずつ右に移動して最初に出る数字を取得
        file_name = None
        current_col = photo_number_col + 1
        while file_name is None and current_col <= sheet.max_column:
            cell_value = sheet.cell(row=photo_number_row, column=current_col).value
            if isinstance(cell_value, (int, float)):  # 数字を確認
                file_name = str(int(cell_value))
            current_col += 1
        
        # ファイル名が見つからない場合はデフォルトのカウンターを使用
        if file_name is None:
            file_name = f'image_{image_counter}'
        
        # 出力ファイル名とパスを生成
        output_image_path = os.path.join(output_folder, f'{file_name}.png')
        
        # 画像を保存
        pil_image.save(output_image_path)
        print(f"Saved {output_image_path}")
        
        image_counter += 1

# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定
resize_to = (900, 600)  # リサイズする幅と高さ（指定しない場合はNone）

extract_images_from_sheet(excel_file, sheet_name, output_folder, resize_to)