from itertools import zip_longest
import re
from django.conf import settings
import openpyxl
from openpyxl.drawing.image import Image
import os

"""
1. クラスを使えば良いのでは？
2. joinの内容はfirstとsecondから1つ目の要素を取り出しているので、メソッドで実現すれば良いのでは？
3. 後に、①などを取り出すのであれば、文字列のインデックス番号( [0] )などを指定して取り出すのではなく、最初から分離しておけば良いのでは？
4. forループ内で、splitなどの文字列操作をしているが、これはdataを作る時点でやっておいたほうが良いのでは？(セルの記入、データの整形を分離させる)
5. 多次元配列にする必要があるのか？
〇6. first second third last はキー名として適切か？ キー名を見ただけでは値が想像できない。
7. damage_coordinate , picture_coordinate は座標であり、xとy の2つを必ず記入する必要がある。故に、要素数不定の配列ではなく、 
   damage_coordinate_x damage_coordinate_y と分離するべきではないか？ 更にキーを短くするため、 damage_x damage_y とするほうが良いと思われる
8. 座標は数値(浮動小数点型)で入力するものであるにも関わらず、文字列型になっているのは何故か？ 
〇9. ①は機種依存文字なので使わない
10. 値が配列になる場合、キー名は複数形であるべきでは？、キー名は名詞単数形もしくは名詞複数形が妥当

"""

# データ (サンプル)
data = [
    {'parts_name': [['排水管 Dp0101']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
     ,
    {'parts_name': [['排水管 Dp0102']], 'damage_name': [['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0102'], 'damage_name': ['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-32', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070486.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['547059.1990495767', '229268.8593029478'], 'picture_coordinate': ['549204.9604817769', '229256.3408485695']}
    ,
    {'parts_name': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'damage_name': [['⑳漏水・滞水-e']], 
     'join': [{'parts_name': ['橋台[胸壁] Ap0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['橋台[竪壁] Ac0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['伸縮装置 Ej0101'], 'damage_name': ['⑳漏水・滞水-e']}], 
     'picture_number': '写真番号-18', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'last_time_picture': 'media\P6100002.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'parts_name': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'damage_name': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'parts_name': ['支承本体 Bh0101'], 'damage_name': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'parts_name': ['沓座モルタル Bm0101'], 'damage_name': ['⑦剥離・鉄筋露出-c']}], 
     'picture_number': '写真番号-27', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'parts_name': [['床版 Ds0201', '床版 Ds0203']], 'damage_name': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'parts_name': ['床版 Ds0201'], 'damage_name': ['⑦剥離・鉄筋露出-d']}, {'parts_name': ['床版 Ds0203'], 'damage_name': ['⑦剥離・鉄筋露出-d']}], 
     'picture_number': None, 'this_time_picture': None, 
     'last_time_picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'parts_name': [['排水管 Dp0103']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
    ,
    {'parts_name': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'damage_name': [['⑳漏水・滞水-e']], 
     'join': [{'parts_name': ['橋台[胸壁] Ap0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['橋台[竪壁] Ac0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['伸縮装置 Ej0101'], 'damage_name': ['⑳漏水・滞水-e']}], 
     'picture_number': '写真番号-18', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'last_time_picture': 'media\P6100020.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'parts_name': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'damage_name': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'parts_name': ['支承本体 Bh0101'], 'damage_name': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'parts_name': ['沓座モルタル Bm0101'], 'damage_name': ['⑦剥離・鉄筋露出-c']}], 
     'picture_number': '写真番号-27', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'parts_name': [['床版 Ds0201', '床版 Ds0203']], 'damage_name': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'parts_name': ['床版 Ds0201'], 'damage_name': ['⑦剥離・鉄筋露出-d']}, {'parts_name': ['床版 Ds0203'], 'damage_name': ['⑦剥離・鉄筋露出-d']}], 
     'picture_number': None, 'this_time_picture': None, 
     'last_time_picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'parts_name': [['排水管 Dp0103']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
    ,
    {'parts_name': [['横桁 Cr0102']], 'damage_name': [['⑰その他(分類6:施工不良)-e']], 
     'join': [{'parts_name': ['横桁 Cr0102'], 'damage_name': ['⑰その他(分類6:施工不良)-e']}], 
     'picture_number': '写真番号-4,5', 'this_time_picture': ['infra/img\\9月7日\u3000佐藤\u3000地上\\P9070424.JPG', 'infra/img\\9月7日\u3000佐藤\u3000地上\\P9070430.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '横桁に施工不良が見られる。', 
     'damage_coordinate': ['532578.7587482664', '229268.8593029478'], 'picture_coordinate': ['532985.6409545547', '228954.2335446222']}
]


# 指定したエクセルファイルを開く
wb = openpyxl.load_workbook('C:\work\django\myproject\program\Infraproject\macro.xlsx')
ws = wb['その１０']  # シート名を指定

# lastがNoneでないデータを残す
filtered_data = [item for item in data if item['this_time_picture'] is not None]

# 行の開始地点と増加するステップを定義
partsname_and_number_row = 10 # 部材名・要素番号
damagename_and_lank_row = 11 # 損傷の種類・損傷程度
picture_start_row = 13 # 損傷写真
lasttime_lank_row = 15 # 前回損傷程度
damage_memo_row = 17 # 損傷メモ
step = 14
num_positions = len(filtered_data) * 3  # データ数に3列分を掛けています

# 関連する列を定義
picture_columns = ["E", "AE", "BE"]
left_columns = ["I", "AI", "BI"]
right_columns = ["R", "AR", "BR"]
bottom_columns = ["T", "AT", "BT"]

# セル位置のリストを生成 ↓
partsname_cell_positions = [] # 部材名
number_cell_positions = [] # 要素番号
damagename_cell_positions = [] # 損傷の種類
lank_cell_positions = [] # 損損傷程度
picture_cell_positions = [] # 損傷写真
lasttime_lank_cell_positions = [] # 前回損傷程度
damage_memo_cell_positions = [] # 損傷メモ

for i in range(num_positions // len(picture_columns)):
    partsname_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in left_columns])
    number_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in right_columns])
    damagename_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in left_columns])
    lank_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in right_columns])
    picture_cell_positions.append([f"{col}{picture_start_row + i * step}" for col in picture_columns])
    lasttime_lank_cell_positions.append([f"{col}{lasttime_lank_row + i * step}" for col in bottom_columns])
    damage_memo_cell_positions.append([f"{col}{damage_memo_row + i * step}" for col in bottom_columns])
    
join_partsname_cell_positions = [item for sublist in partsname_cell_positions for item in sublist]
join_number_cell_positions = [item for sublist in number_cell_positions for item in sublist]
join_damagename_cell_positions = [item for sublist in damagename_cell_positions for item in sublist]
join_lank_cell_positions = [item for sublist in lank_cell_positions for item in sublist]
join_picture_cell_positions = [item for sublist in picture_cell_positions for item in sublist]
join_lasttime_lank_cell_positions = [item for sublist in lasttime_lank_cell_positions for item in sublist]
join_damage_memo_cell_positions = [item for sublist in damage_memo_cell_positions for item in sublist]
# セル位置のリストを生成 ↑

number_change = {
'①': '腐食',
'②': '亀裂',
'③': 'ゆるみ・脱落',
'④': '破断',
'⑤': '防食機能の劣化',
'⑥': 'ひびわれ',
'⑦': '剥離・鉄筋露出',
'⑧': '漏水・遊離石灰',
'⑨': '抜け落ち',
'⑩': '補修・補強材の損傷',
'⑪': '床版ひびわれ',
'⑫': 'うき',
'⑬': '遊間の異常',
'⑭': '路面の凹凸',
'⑮': '舗装の異常',
'⑯': '支承部の機能障害',
'⑰': 'その他',
'⑱': '定着部の異常',
'⑲': '変色・劣化',
'⑳': '漏水・滞水',
'㉑': '異常な音・振動',
'㉒': '異常なたわみ',
'㉓': '変形・欠損',
'㉔': '土砂詰まり',
'㉕': '沈下・移動・傾斜',
'㉖': '洗掘',
}

# 最大の写真サイズ（幅、高さ）
max_width, max_height = 240, 180 # 4:3

# 位置を追跡するカウンタ
cell_counter = 0

for item in filtered_data:
    # 3列目(インデックスが2)でpictureキーに写真が設定されている場合
    if cell_counter % 2 == 0 and item['last_time_picture'] is not None:
        # 3列目を空白にするため、インデックスを1つ追加
        cell_counter += 1
    # pictureキーに写真が設定されていても、3列目でなければOK
    else:
        pass

    # lastの写真を張る動作
    for image_path in item['this_time_picture']:
        if os.path.exists(image_path):
            img = Image(image_path) # 画像の読み込み
            img.width, img.height = max_width, max_height # 画像サイズの設定    
            cell_pos = picture_cell_positions[cell_counter // len(picture_columns)][cell_counter % len(picture_columns)] # 所定のセル位置       
            ws.add_image(img, cell_pos) # ワークシートに画像を追加 
            cell_counter += 1 # カウンタを一つ進める
            
    # pictureの写真を張る動作
    if item['last_time_picture'] and os.path.exists(item['last_time_picture']):
        img = Image(item['last_time_picture'])
        img.width, img.height = max_width, max_height
        cell_pos = picture_cell_positions[cell_counter // len(picture_columns)][cell_counter % len(picture_columns)]
        ws.add_image(img, cell_pos)
        cell_counter += 1

# データの入力
data_index = 0

for item, part_pos, number_pos, name_pos, lank_pos, memo_pos, picture_pos in zip_longest(filtered_data, join_partsname_cell_positions, join_number_cell_positions, join_damagename_cell_positions, join_lank_cell_positions, join_damage_memo_cell_positions, join_picture_cell_positions, fillvalue=None):
    
    if (data_index == 2 or data_index % 2 == 3) and item['last_time_picture'] is not None:
        # 3列目を空白にするため、インデックスを1つ追加
        data_index += 1
        
    part_cell = join_partsname_cell_positions[data_index]
    number_cell = join_number_cell_positions[data_index]
    name_cell = join_damagename_cell_positions[data_index]
    post_lank_cell = join_lank_cell_positions[data_index]
    pre_lank_cell = join_lasttime_lank_cell_positions[data_index]
    memo_cell = join_damage_memo_cell_positions[data_index]
    
# メモに入れるための固定コード　↓
    # firstキーの内容を所定の書式に変更
    try:
        first_data = item['parts_name'][0][0] # 例: "排水管 Dp0101"
        split_space = first_data.split(" ")
        first_part_data = split_space[0]
        
        match = re.search(r'\d+', split_space[1])
        first_number_data = match.group() if match else ''
        
        ws[part_cell] = first_part_data # 例: "排水管"
        ws[number_cell] = first_number_data # 例: "0101"
        
        # secondキーの内容を所定の書式に変更
        second_data = item['damage_name'][0][0] # 例: "①腐食(大大)-e"
        second_damage_name = second_data[0] # 例: "①"
        second_name_data = number_change.get(second_damage_name, second_damage_name)
        second_lank_data = second_data[-1]
        
        ws[name_cell] = second_name_data # 例: "腐食"
        ws[post_lank_cell] = second_lank_data # 例: "e"
        
        # textarea_contentキーの内容を所定の書式に変更
        memo_data = item['textarea_content'] # 例: "排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e"
        ws[memo_cell] = memo_data
            
        if item['last_time_picture'] is not None:
            data_index += 2
        else:
            data_index += 1
    except (TypeError, KeyError):
        ws[part_cell] = ""
        ws[number_cell] = ""
        ws[name_cell] = ""
        ws[post_lank_cell] = ""
        ws[memo_cell] = ""
# メモに入れるための固定コード　↑
        
# エクセルファイルを保存
# 別名で保存するため、新しいファイル名を指定
if os.name == 'nt':  # Windowsの場合
    new_file_path = 'C:/Users/dobokuka4/Desktop/tameshi.xlsx'
elif os.name == 'posix':  # Macの場合 (および他のUNIX系)
    new_file_path = '/Users/YourUsername/Desktop/example.xlsx'
else:
    raise EnvironmentError('Unsupported platform')

# 新しいファイル名で保存
wb.save(new_file_path)
