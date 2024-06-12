import os
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO

def extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # 「写真番号」と書かれているセルを探す
    photo_number_positions = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "写真番号":
                photo_number_positions.append((cell.row, cell.column))
                #print(f"写真番号 Row: {cell.row}, Column: {cell.column}")

    # 写真番号セルから右方向に移動した位置の値を取得するための辞書
    photo_name_dict = {}
    for p_row, p_col in photo_number_positions:
        for offset in range(1, 10):  # 最大10列までチェックする
            photo_name_cell = sheet.cell(row=p_row, column=p_col + offset)
            if photo_name_cell.value:
                photo_name_dict[(p_row, p_col)] = photo_name_cell.value
                #print(f"Photo name at ({p_row}, {p_col}): {photo_name_cell.value}")
                break

    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))

    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 写真番号セルから右方向の値で画像を保存
    used_photo_names = set()  # 既に使用された写真番号のセット
    
    for row, col, img in images:
        photo_name = None
        for (p_row, p_col), name in photo_name_dict.items():
            if row == p_row and col > p_col:
                photo_name = name
                break
        
        if not photo_name:
            # 指定された位置に写真番号がない場合、デフォルトの名前を使用
            photo_name = f"photo_{row}_{col}"
        
        # 重複しないファイル名を生成
        original_photo_name = photo_name
        count = 1
        while photo_name in used_photo_names:
            photo_name = f"{original_photo_name}_{count}"
            count += 1
        used_photo_names.add(photo_name)

        # 画像を保存する
        img_data = img._data()
        #image_file_path = os.path.join(output_folder, f"{photo_name}.jpg")
        # 各写真のファイル名をループで生成
        # ここでphoto_nameは辞書の値を取り出している
        image_file_path = os.path.join(output_folder, f"{photo_name}.jpg")
        
        with open(image_file_path, "wb") as img_file:
            img_file.write(img_data)
            print(img_file)


# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定

extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder)