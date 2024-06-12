import re
import openpyxl
from openpyxl.drawing.image import Image
import os

# データ (サンプル)
data = [
    {'first': [['排水管 Dp0101']], 'second': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'first': ['排水管 Dp0101'], 'second': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'third': '写真番号-31', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
     ,
    {'first': [['排水管 Dp0102']], 'second': [['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'first': ['排水管 Dp0102'], 'second': ['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']}], 
     'third': '写真番号-32', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070486.JPG'], 
     'picture': None, 
     'textarea_content': '排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['547059.1990495767', '229268.8593029478'], 'picture_coordinate': ['549204.9604817769', '229256.3408485695']}
    ,
    {'first': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'second': [['⑳漏水・滞水-e']], 
     'join': [{'first': ['橋台[胸壁] Ap0101'], 'second': ['⑳漏水・滞水-e']}, {'first': ['橋台[竪壁] Ac0101'], 'second': ['⑳漏水・滞水-e']}, {'first': ['伸縮装置 Ej0101'], 'second': ['⑳漏水・滞水-e']}], 
     'third': '写真番号-18', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'picture': 'C:\work\django\myproject\program\Infraproject\media\P6100002.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'first': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'second': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'first': ['支承本体 Bh0101'], 'second': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'first': ['沓座モルタル Bm0101'], 'second': ['⑦剥離・鉄筋露出-c']}], 
     'third': '写真番号-27', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'first': [['床版 Ds0201', '床版 Ds0203']], 'second': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'first': ['床版 Ds0201'], 'second': ['⑦剥離・鉄筋露出-d']}, {'first': ['床版 Ds0203'], 'second': ['⑦剥離・鉄筋露出-d']}], 
     'third': None, 'last': None, 
     'picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'first': [['排水管 Dp0103']], 'second': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'first': ['排水管 Dp0101'], 'second': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'third': '写真番号-31', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
     ,
    {'first': [['排水管 Dp0102']], 'second': [['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'first': ['排水管 Dp0102'], 'second': ['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']}], 
     'third': '写真番号-32', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070486.JPG'], 
     'picture': None, 
     'textarea_content': '排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['547059.1990495767', '229268.8593029478'], 'picture_coordinate': ['549204.9604817769', '229256.3408485695']}
    ,
    {'first': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'second': [['⑳漏水・滞水-e']], 
     'join': [{'first': ['橋台[胸壁] Ap0101'], 'second': ['⑳漏水・滞水-e']}, {'first': ['橋台[竪壁] Ac0101'], 'second': ['⑳漏水・滞水-e']}, {'first': ['伸縮装置 Ej0101'], 'second': ['⑳漏水・滞水-e']}], 
     'third': '写真番号-18', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'picture': 'C:\work\django\myproject\program\Infraproject\media\P6100002.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'first': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'second': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'first': ['支承本体 Bh0101'], 'second': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'first': ['沓座モルタル Bm0101'], 'second': ['⑦剥離・鉄筋露出-c']}], 
     'third': '写真番号-27', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'first': [['床版 Ds0201', '床版 Ds0203']], 'second': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'first': ['床版 Ds0201'], 'second': ['⑦剥離・鉄筋露出-d']}, {'first': ['床版 Ds0203'], 'second': ['⑦剥離・鉄筋露出-d']}], 
     'third': None, 'last': None, 
     'picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'first': [['排水管 Dp0103']], 'second': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'first': ['排水管 Dp0101'], 'second': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'third': '写真番号-31', 'last': ['C:\work\django\myproject\program\Infraproject\infra\static\infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
]

# 指定したエクセルファイルを開く
import re
import openpyxl
from openpyxl.drawing.image import Image
import os

# データ (サンプル) - ここは省略

# 指定したエクセルファイルを開く
wb = openpyxl.load_workbook('C:\work\django\myproject\program\Infraproject\macro.xlsx')
ws = wb['その１０']  # シート名を指定

# lastがNoneでないデータを残す
filtered_data = [item for item in data if item['last'] is not None]

# 行の開始地点と増加するステップを定義
partsname_and_number_row = 10 # 部材名・要素番号
damagename_and_lank_row = 11 # 損傷の種類・損傷程度
picture_start_row = 13 # 損傷写真
lasttime_lank_row = 15 # 前回損傷程度
damage_memo_row = 17 # 損傷メモ
step = 14
num_positions = len(filtered_data) * 3  # データ数に3列分を掛けています

# 関連する列を定義
picture_columns = ["E", "AE", "BE"]
left_columns = ["I", "AI", "BI"]
right_columns = ["R", "AR", "BR"]
bottom_columns = ["T", "AT", "BT"]

# セル位置のリストを動的に生成
partsname_cell_positions = [] # 部材名
number_cell_positions = [] # 要素番号
damagename_cell_positions = [] # 損傷の種類
lank_cell_positions = [] # 損損傷程度
picture_cell_positions = [] # 損傷写真
lasttime_lank_cell_positions = [] # 前回損傷程度
damage_memo_cell_positions = [] # 損傷メモ

for i in range(num_positions // len(picture_columns)):
    partsname_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in left_columns])
    number_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in right_columns])
    damagename_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in left_columns])
    lank_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in right_columns])
    picture_cell_positions.append([f"{col}{picture_start_row + i * step}" for col in picture_columns])
    lasttime_lank_cell_positions.append([f"{col}{lasttime_lank_row + i * step}" for col in bottom_columns])
    damage_memo_cell_positions.append([f"{col}{damage_memo_row + i * step}" for col in bottom_columns])
                    
number_change = {
'①': 'A',
'②': 'B',
'③': 'C',
}

# データの入力
data_index = 0

for part_pos, number_pos, name_pos, lank_pos, memo_pos, picture_pos in zip(partsname_cell_positions, number_cell_positions, damagename_cell_positions, lank_cell_positions, damage_memo_cell_positions, picture_cell_positions):
    for idx, (part_cell, number_cell, name_cell, lank_cell, memo_cell, pic_cell) in enumerate(zip(part_pos, number_pos, name_pos, lank_pos, memo_pos, picture_pos)):
        # データ範囲内であれば処理を行う
        if data_index >= len(filtered_data):
            break

        current_data = filtered_data[data_index]

        # lastキーがNoneの場合、データをスキップ
        if current_data['last'] is None:
            data_index += 1
            continue

        picture_path = current_data['picture']
        if idx == 2 and picture_path is not None:
            continue

        # データの入力処理
        first_data = current_data.get('first', [['']])[0][0] # 例: "排水管 Dp0101"
        split_space = first_data.split(" ")
        first_part_data = split_space[0]
        
        match = re.search(r'\d+', split_space[1])
        first_number_data = match.group() if match else ''
        
        ws[part_cell] = first_part_data # 例: "排水管"
        ws[number_cell] = first_number_data # 例: "0101"
        
        # secondキーの内容を右側のセルに入力
        second_data = current_data.get('second', [['']])[0][0] # 例: "①腐食(大大)-e"
        second_damage_name = second_data[0] # 例: "①"
        second_name_data = number_change.get(second_damage_name, second_damage_name)
        second_lank_data = second_data[-1]
        
        ws[name_cell] = second_name_data # 例: "腐食"
        ws[lank_cell] = second_lank_data # 例: "e"
        
        # memoデータを右側のセルに入力
        memo_data = current_data.get('textarea_content', '') # 例: "排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e"
        ws[memo_cell] = memo_data
        
        data_index += 1
            
# エクセルファイルを保存
# 別名で保存するため、新しいファイル名を指定
new_file_path = 'C:/Users/dobokuka4/Desktop/tameshi.xlsx'
# macの場合(new_file_path = '/Users/YourUsername/Desktop/example.xlsx')

# 新しいファイル名で保存
wb.save(new_file_path)
