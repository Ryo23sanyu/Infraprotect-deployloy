# アプリ内からインポート
from collections import defaultdict
import datetime
from functools import reduce
from io import BytesIO
from itertools import groupby, zip_longest
import json
import logging
import math
from operator import attrgetter
import re
import os
import glob
from django.db import IntegrityError
import openpyxl
import tempfile
# サードパーティー製モジュール
import ezdxf
import pandas as pd
import urllib.parse
from markupsafe import Markup
from copy import copy
from itertools import zip_longest
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
# django内からインポート
from django.http import FileResponse, Http404, HttpResponse, HttpResponseBadRequest, HttpResponseNotFound, JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.files.storage import FileSystemStorage
from django.views.generic import ListView, DetailView, CreateView, DeleteView, UpdateView
from django.views.decorators.csrf import csrf_protect, csrf_exempt

from infraproject import settings
from .models import Approach, Article, DamageComment, DamageList, DamageReport, FullReportData, Infra, PartsName, PartsNumber, Table, LoadGrade, LoadWeight, Photo, Panorama, NameEntry, Regulation, Rulebook, Thirdparty, UnderCondition, Material
from .forms import BridgeCreateForm, BridgeUpdateForm, CensusForm, DamageCommentCauseEditForm, DamageCommentEditForm, DamageCommentJadgementEditForm, EditReportDataForm, FileUploadForm, FullReportDataEditForm, NameEntryForm, PartsNumberForm, TableForm, UploadForm, PhotoUploadForm, NameForm
from .forms import ArticleForm
from urllib.parse import quote, unquote

class ListInfraView(LoginRequiredMixin, ListView):
    template_name = 'infra/infra_list.html'
    model = Infra # 使用するモデル「infra」
    def get_queryset(self, **kwargs):
        # モデル検索のクエリー。Infra.objects.all() と同じ結果で全ての Infra
        queryset = super().get_queryset(**kwargs)
        # パスパラメータpkによりarticleを求める
        # 指定されたpk(idの指定)のデータを取得
# article  = Article.objects.get(id = self.kwargs["pk"])
        # get使用すると、存在しない場合エラーになってしまう
        # 求めたarticleを元にモデル検索のクエリーを絞り込む
        # infra_objectフィルタ－
        #queryset = queryset.filter(article=article)
        queryset = queryset.filter(article = self.kwargs["article_pk"])
        # 絞り込んだクエリーをDjangoに返却し表示データとしてもらう
        return queryset
    def get_context_data(self, **kwargs):
        # HTMLテンプレートでの表示変数として「article_id」を追加。
        # 値はパスパラメータpkの値→取り扱うarticle.idとなる
        kwargs["article_id"] = self.kwargs["article_pk"]
        return super().get_context_data(**kwargs)


class DetailInfraView(LoginRequiredMixin, DetailView):
    template_name = 'infra/infra_detail.html'
    model = Infra
    def get_context_data(self, **kwargs):
        # HTMLテンプレートでの表示変数として「article_id」を追加。
        # 値はパスパラメータpkの値→取り扱うarticle.idとなる
        kwargs["article_id"] = self.kwargs["article_pk"]
        #モデルのTableクラス ↑                    ↑  infraに格納する値は自らのpkの値とする
        return super().get_context_data(**kwargs)
  
class CreateInfraView(LoginRequiredMixin, CreateView):
    template_name = 'infra/infra_create.html' # 対応するhtmlの名前
    model = Infra # models.pyのどのモデルと紐付くか
    # form_class = BridgeCreateForm # forms.pyのどのクラスと紐付くか
    fields = ('title', '径間数', '橋長', '全幅員','橋梁コード', '活荷重', '等級', '適用示方書', 
              '上部構造形式', '下部構造形式', '基礎構造形式', '近接方法', '交通規制', '第三者点検', '海岸線との距離', 
              '路下条件', '交通量', '大型車混入率', '特記事項', 'カテゴリー', 'latitude', 'longitude', 'article')
    
    def form_valid(self, form):
        article_pk = self.kwargs['article_pk']
        print(article_pk)
        article = get_object_or_404(Article, pk=article_pk)
        print(article)
        form.instance.article = article
        print(form.instance)
        return super().form_valid(form)

    def get_success_url(self):
        # Define where you want to redirect after successful form submission
        return reverse('article-detail', kwargs={'pk': self.object.article.pk})

    #新規作成時、交通規制の全データをコンテキストに含める。
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["loadWeights"] = LoadWeight.objects.all()
        context["loadGrades"] = LoadGrade.objects.all()
        context["rulebooks"] = Rulebook.objects.all()
        context["approachs"] = Approach.objects.all()
        context["regulations"] = Regulation.objects.all()
        context["thirdpartys"] = Thirdparty.objects.all()
        context["underconditions"] = UnderCondition.objects.all()
        return context
        
    def keikan_create_view(request, article_pk, pk): # keikan_create_view関数を定義
        if request.method == "POST": # リクエストメソッドがpostの場合
            form = BridgeCreateForm(request.POST) # BridgeCreateFormというフォームクラスのインスタンスを生成
            if form.is_valid(): # formが正常の場合
                keikan_number = form.cleaned_data['径間数'] # form.cleaned_dataから'径間数'キーに対応するデータを取得
                request.session['keikan_number'] = keikan_number # 取得した"径間数"を現在のユーザーセッションに保存
                # print(form.instance.pk)
                return redirect('bridge-table', article_pk, pk) # 「table」という名前のURLにリダイレクト
            else:
                form = BridgeCreateForm() # 新しい空のフォームインスタンスを生成
        return render(request, 'infra/infra_create.html', {'form': form, 'object': Table.objects.filter(id=pk).first()})
        # 'infra_create.html'テンプレートをレンダリング
    
    def damage_view(request, article_pk, pk): # damage_view関数を定義
        keikan_number = request.session.get('keikan_number', 1) # request.session.getメソッドを使い、セッションから"径間数"を取得、デフォルト値は1
        keikan_range = list(range(keikan_number)) # 1からkeikan_number（"径間数"）までの連続する整数列を生成
        return render(request, 'bridge_table.html', {'keikan_range': keikan_range, 'object': Table.objects.filter(id=pk).first()})
        # 'table.html'テンプレートをレンダリング

class DeleteInfraView(LoginRequiredMixin, DeleteView):
    template_name = 'infra/infra_delete.html'
    model = Infra
    success_url = reverse_lazy('list-infra')
    def get_success_url(self):
        return reverse_lazy('list-infra', kwargs={'article_pk': self.kwargs["article_pk"]})
      
class UpdateInfraView(LoginRequiredMixin, UpdateView):
    template_name = 'infra/infra_update.html'
    model = Infra
    fields = ('title', '径間数', '橋長', '全幅員', 'latitude', 'longitude', '橋梁コード', '活荷重', '等級', '適用示方書', 
              '上部構造形式', '下部構造形式', '基礎構造形式', '近接方法', '交通規制', '第三者点検', '海岸線との距離', 
              '路下条件', '交通量', '大型車混入率', '特記事項', 'カテゴリー', 'article')
    success_url = reverse_lazy('detail-infra')
    def get_success_url(self):
        return reverse_lazy('detail-infra', kwargs={'article_pk': self.kwargs["article_pk"], 'pk': self.kwargs["pk"]})

    #新規作成時、交通規制の全データをコンテキストに含める。
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 編集中のインスタンスに紐づく交通規制のIDをリストとして取得
        # コンテキストに追加
        selected_regulations = self.object.交通規制.values_list('id', flat=True)# 選択状態を保持
        context['selected_regulations'] = list(selected_regulations)# 選択状態を保持
        context["regulations"] = Regulation.objects.all()
        
        selected_loadWeights = self.object.活荷重.values_list('id', flat=True)
        context['selected_loadWeights'] = list(selected_loadWeights)
        context["loadWeights"] = LoadWeight.objects.all()
        
        selected_loadGrades = self.object.等級.values_list('id', flat=True)
        context['selected_loadGrades'] = list(selected_loadGrades)
        context["loadGrades"] = LoadGrade.objects.all()
        
        selected_rulebooks = self.object.適用示方書.values_list('id', flat=True)
        context['selected_rulebooks'] = list(selected_rulebooks)
        context["rulebooks"] = Rulebook.objects.all()
        
        selected_approachs = self.object.近接方法.values_list('id', flat=True)
        context['selected_approachs'] = list(selected_approachs)
        context["approachs"] = Approach.objects.all()
        
        selected_thirdpartys = self.object.第三者点検.values_list('id', flat=True)
        context['selected_thirdpartys'] = list(selected_thirdpartys)
        context["thirdpartys"] = Thirdparty.objects.all()
        
        selected_underconditions = self.object.路下条件.values_list('id', flat=True)
        context['selected_underconditions'] = list(selected_underconditions)
        context["underconditions"] = UnderCondition.objects.all()
        return context

def infra_view(request):
    if request.method == 'POST':
        等級 = request.POST.get('等級', None)
        # load_gradeを使って必要な処理を行う
        # 例えば、選択されたload_gradeに基づいてデータをフィルタリングして表示するなど

    # 通常のビューロジック
    return render(request, 'infra/infra_detail.html')

def index_view(request):
    order_by = request.GET.get('order_by', '案件名')
    object_list = Article.objects.order_by(order_by)
    return render(request, 'infra/index.html', {'object_list': object_list})

class ListArticleView(LoginRequiredMixin, ListView):
    template_name = 'infra/article_list.html'
    model = Article
  
class DetailArticleView(LoginRequiredMixin, DetailView):
    template_name = 'infra/article_detail.html'
    model = Article
    
# この関数はファイルツリーを生成する
class CreateArticleView(LoginRequiredMixin, CreateView):
    template_name = 'infra/article_create.html'
    model = Article
    fields = ('案件名', '土木事務所', '対象数', '担当者名', 'その他', 'ファイルパス')
    success_url = reverse_lazy('list-article')

    def get_initial(self):
        initial = super().get_initial()
        selected_file = self.request.COOKIES.get('selected_file')
        if selected_file:
            initial['ファイルパス'] = selected_file
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # root_dir = os.path.expanduser('~') # Cドライブのユーザーディレクトリ
        user_home_dir = os.path.expanduser('~')
        root_dir = os.path.join(user_home_dir, 'Desktop') # デスクトップディレクトリ
        context['root_dir'] = root_dir
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        response.delete_cookie('selected_file')
        return response

def get_subdirectories(request):
    # | をバックスラッシュに戻します
    path = unquote(request.GET.get('path', ''))
    if not path:
        return JsonResponse({'directories': [], 'files': []}, status=400)

    subdirectories = []
    files = []

    try:
        with os.scandir(path) as it:
            for entry in it:
                if entry.is_dir():
                    subdirectories.append(entry.name)
                elif entry.is_file():
                    files.append(entry.name)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'directories': subdirectories, 'files': files})
    
class DeleteArticleView(LoginRequiredMixin, DeleteView):
    template_name = 'infra/article_delete.html'
    model = Article
    success_url = reverse_lazy('list-article')
  
class UpdateArticleView(LoginRequiredMixin, UpdateView):
    template_name = 'infra/article_update.html'
    model = Article
    fields = ('案件名', '土木事務所', '対象数', '担当者名', 'その他', 'ファイルパス')
    success_url = reverse_lazy('list-article')

# << ファイルのアップロード・各infraに紐付け >>
def file_upload(request, article_pk, pk):
    if request.method == 'POST':
                #                  ↓  request.POST の中にdxfファイルの名前が入っているだけ。.copy() を実行して編集可能にする。
        copied          = request.POST.copy()

        # ここで Infraのid(pk)を指定する。
        copied["infra"] = pk
        copied["article"] = article_pk
        
        # バリデーション
        form = TableForm(copied, request.FILES)
        
        # 既存のオブジェクトに対して新しいファイルを上書きする処理
        if Table.objects.filter(infra=pk).first():
            obj = Table.objects.get(infra=pk)
            form = TableForm(copied, request.FILES, instance=obj)

        if form.is_valid():
            form.save()
            return redirect('file_upload_success')
    else:
        form = TableForm()
    
    return render(request, 'infra/file_upload.html', {'form': form, 'article_pk': article_pk, 'pk': pk})

def file_upload_success(request):
    return render(request, 'infra/file_upload_success.html')
  
def photo_list(request):
    photos = Photo.objects.all()
    return render(request, 'infra/photo_list.html', {'photos': photos})

def selected_photos(request):
    selected_photo_ids = request.POST.getlist('selected_photos')
    selected_photos = Photo.objects.filter(id__in=selected_photo_ids)
    return render(request, 'infra/selected_photos.html', {'selected_photos': selected_photos})


def panorama_list(request):
    panoramas = Panorama.objects.all()
    if request.method == 'POST':
        selected_ids = request.POST.getlist('image_list')
        for panorama in panoramas:
            if str(panorama.id) in selected_ids:
                panorama.checked = True
            else:
                panorama.checked = False
            panorama.save()
        return redirect('image_list')  # 再描画のためにリダイレクト
    
    return redirect('image_list')
    #return render(request, 'image_list.html', {'panoramas': panoramas})

def panorama_upload(request):
    if request.method == 'POST':
        image = request.FILES['image']
        checked = request.POST.get('checked', False)
        panorama = Panorama.objects.create(image=image, checked=checked)
        return redirect('photo')
    return render(request, 'panorama_upload.html')

# << センサス調査 >>
def census_view(request):
    form = CensusForm()
    return render(request, 'infra_detail.html', {'form': form})

# << 全景写真の表示 >>
def image_list(request):
    # 特定のディレクトリ内の全てのファイルパスをリストで取得したい場合はglobを使うと良い。
    save_path = str(settings.BASE_DIR) + "\infra\static\infra\img"
    # 「C:\work\django\myproject\program\Infraproject」+「\infra\static\infra\img」と同意
    files_jpg = glob.glob(save_path + "\*.jpg")
    files_png = glob.glob(save_path + "\*.png")
    files = files_jpg + files_png  # 2つのリストを結合する。
    # ページに表示する際、"infra/static/" を削除する。

    web_base_path = "infra/img/"  # ウェブからアクセスする際のベースパス

    image_files = []
    for file_path in files:
        # OSの絶対パスからウェブアクセス可能な相対パスへ変換
        relative_path = file_path.replace(save_path, web_base_path).replace("\\", "/")
        image_files.append(relative_path)
    # テンプレートに画像ファイルの一覧を渡してレンダリングする
    return render(request, 'image_list.html', {'image_files': image_files})

# << 全景写真アップロード >>
save_path = str(settings.BASE_DIR) + "\infra\static\infra\img"
# 「C:\work\django\myproject\program\Infraproject」+「\infra\static\infra\img」と同意
def display_photo(request):
    print("リクエストメソッド:", request.method)  # リクエストのメソッドを表示
    if request.method == 'POST': # HTTPリクエストがPOSTメソッド(フォームの送信など)であれば、以下のコードを実行
        form = UploadForm(request.POST, request.FILES) # UploadFormを使用して、送信されたデータ(request.POST)とファイル(request.FILES)を取得
        print("フォームが有効か？:", form.is_valid())  # フォームの有効性を表示
        if form.is_valid(): # formのis_valid()メソッドを呼び出して、フォームのバリデーション(検証)を実行
    # フォームが有効な場合は、選択された写真を特定のフォルダに保存します
            # print("フォームが有効")
            photo = form.cleaned_data['photo'] # バリデーションを通過したデータから、'photo'キーに対応するデータを取得しphoto変数に格納
            print("アップロードされた写真の名前:", photo.name)  # アップロードされた写真の名前を表示
            file_name = photo.name  # アップロードされたファイルの名前をfile_name変数に格納
            file_path = os.path.join(save_path, file_name)
            #設定された保存パス（save_path）とファイル名（file_name）を組み合わせ、フルパスをfile_path変数に格納

            with open(file_path, 'wb') as f: # file_pathで指定されたパスにファイルをバイナリ書き込みモード('wb')で開く
                for chunk in photo.chunks(): # アップロードされたファイル(photo)をchunks()メソッドを使用して分割し、ループ処理を行う
                    f.write(chunk) # 各チャンクを開かれたファイル(f)に書き込む
            return redirect('image_list')
        else:
            print("フォームエラー:", form.errors)  # フォームのエラーを表示
        # フォームが有効・無効 共通
            print(form.errors)  # コンソールにエラーメッセージを出力
            return HttpResponseBadRequest('添付ファイルが見つかりませんでした。 エラー: {}'.format(form.errors))
            #return HttpResponseBadRequest('添付ファイルが見つかりませんでした。')
    else:
        form = UploadForm() # Forms.pyの「UploadForm」を呼び出し
    return render(request, 'upload_photo.html', {'form': form})

def photo_upload(request):
    if request.method == 'POST':
        form = PhotoUploadForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('image_list')
    else:
        form = PhotoUploadForm()
    return render(request, 'image_list.html', {'form': form})

# << 全景写真の変更 >>
def change_photo(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            # ここで、古い写真の削除や新しい写真の保存処理を行います。
            # 例: image_idを使用して特定のレコードを探し、そのレコードに新しい画像を保存
            image_id = request.POST.get('image_id', None)  # 画像ID取得
            if image_id:
                image_instance = UploadForm.objects.get(id=image_id)
                image_instance.photo = form.cleaned_data['photo']
                image_instance.save()
                return redirect('image_list')
            else:
                # エラー処理
                pass
        else:
            # フォームのバリデーションに失敗したときの処理
            pass
    else:
        form = UploadForm()  # GETリクエストの場合、空のフォームを表示

    return render(request, 'upload_photo.html', {'form': form})

# << 損傷写真帳の作成 >>
def bridge_table(request, article_pk, pk): # idの紐付け infra/bridge_table.htmlに表示
    context = {}
    # プロジェクトのメディアディレクトリからdxfファイルまでの相対パス
    # URL：article/<int:article_pk>/infra/<int:pk>/bridge-table/
    table = Table.objects.filter(infra=pk).first()
    print(f"table_name:{table}") # Table object (17)
    print(f"table.infra.title:{table.infra.title}")
    infra_instance = Infra.objects.filter(title = table.infra.title)
    print(f"table_name:{infra_instance}") # センサス橋
    # オブジェクトディレクトリの相対パスを取得
    if table.dxf:
        encoded_url_path = table.dxf.url
    decoded_url_path = urllib.parse.unquote(encoded_url_path) # URLデコード
    # 絶対パスと合体
    dxf_filename = os.path.join(settings.BASE_DIR, decoded_url_path.lstrip('/'))
    print(f"dxfファイルパス:{table.dxf.url}")
    # context["object"] = table
    # keikan_infra = Infra.objects.filter(id=pk).first() # 271行目と同じ
    # context["buttons"] = table.infra.径間数 * " " # Tableクラスのinfraオブジェクトから「径間数」を取り出す
    
    # bridge_tableのボタンを押したときのアクション

    second_search_title_text = "損傷図"
    
    # sorted_items = create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text)

    # << 辞書型として、全径間を1つの多重リストに格納 >>
    max_search_title_text = table.infra.径間数
    database_sorted_items = []  # 結果をまとめるリスト
    
    for search_title_text_with_suffix in range(1, max_search_title_text + 1):
        search_title_text = f"{search_title_text_with_suffix}径間"
        sub_database_sorted_items = create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text)
        for item in sub_database_sorted_items:
            item['search'] = search_title_text
            database_sorted_items.append(item)
    # print(f"database_sorted_items:{database_sorted_items}")
    """辞書型の多重リストをデータベースに登録"""
    # << ['']を外してフラットにする >>
    def flatten(value):
        def _flatten(nested_list):
            if isinstance(nested_list, list):
                for item in nested_list:
                    yield from _flatten(item)
            else:
                yield nested_list
        
        return ', '.join(_flatten(value))

    # << joinキーを変換 >>
    def join_to_result_string(join):
        result_parts = []
        for item in join:
            parts_name = item['parts_name'][0]
            damage_names = item['damage_name']
            formatted_damage_names = '/'.join(damage_names)
            result_parts.append(f"{parts_name} : {formatted_damage_names}")
        return ', '.join(result_parts)

    # << 写真のキーを変換 >>
    def simple_flatten(value):
        return ', '.join(map(str, value)) if isinstance(value, list) else value
    
    # <<正規表現で4桁以上の番号を取得>>
    def extract_number(text):
        pattern = r'\d{4,}' # 4文字以上の連続する数字
        matches = re.findall(pattern, text)
        return matches
    
    for damage_data in database_sorted_items:
        #print(f"damage_data：{damage_data}")
        # 元の辞書から 'picture_number' の値を取得
        #             　辞書型 ↓           ↓ キーの名前      ↓ 存在しない場合、デフォルト値として空白を返す
        picture_number = damage_data.get('picture_number', '')
        # 正規表現で数字のみを抽出
        if picture_number:
            # 数字のみを抽出
            numbers_only = re.findall(r'\d+', str(picture_number)) # ['2']
            # 数字を結合して整数に変換
            numbers_only = int(''.join(numbers_only)) if numbers_only else None # 2
        else:
            numbers_only = None

        damage_coordinate = damage_data.get('damage_coordinate', [None, None])
        picture_coordinate = damage_data.get('picture_coordinate', [None, None])

        #parts_list = flatten(damage_data.get('parts_name', ''))
        #damage_list = flatten(damage_data.get('damage_name', ''))

        names = damage_data.get('parts_name', '')
        damages = damage_data.get('damage_name', '')
        #print(f"names:{names}")
        #print(f"damages:{damages}")
        
        split_names = []

        for item in names:
            split_items = []
            for split in item:
                if "～" in split:
                    one = split.find("～")
                    start_number = ''.join(extract_number(split[:one])) # 0101
                    end_number = ''.join(extract_number(split[one+1:])) # 0204

                    # 最初の2桁と最後の2桁を取得
                    start_prefix = start_number[:2] # 01
                    start_suffix = start_number[2:] # 01
                    end_prefix = end_number[:2] # 01
                    end_suffix = end_number[2:] # 03
                    
                    part_name = split[:one].replace(start_number, '')
                
                    for prefix in range(int(start_prefix), int(end_prefix)+1):
                        for suffix in range(int(start_suffix), int(end_suffix)+1):
                            number_items = "{:02d}{:02d}".format(prefix, suffix)
                            join_item = part_name + number_items
                            split_items.append(join_item)
                    
                else:
                    split_items.append(split)
            split_names.append(split_items)
        
        join = join_to_result_string(damage_data.get('join', ''))
        this_time_picture = simple_flatten(damage_data.get('this_time_picture', ''))
        last_time_picture = simple_flatten(damage_data.get('last_time_picture', ''))
        textarea_content = damage_data.get('textarea_content', '')
        damage_coordinate_x = damage_coordinate[0] if damage_coordinate else None
        damage_coordinate_y = damage_coordinate[1] if damage_coordinate else None
        picture_coordinate_x = picture_coordinate[0] if picture_coordinate else None
        picture_coordinate_y = picture_coordinate[1] if picture_coordinate else None
        span_number = damage_data.get('search', '')
        print("----------------------------------")
        print(f"damage_data:{damage_data}")
        print(f"this_time_picture:{this_time_picture}")
        name_length = len(split_names)
        damage_length = len(damages)
        
        # 多重リストかどうかを判定する関数
        def is_multi_list(lst):
            return any(isinstance(i, list) for i in lst)
        
        def process_names(names):
            """
            与えられたnamesを処理し、適切な部分を返す関数。
            """
            
            parts_left = ["主桁", "PC定着部"]  # 左の数字
            parts_right = ["横桁", "橋台"]     # 右の数字
            parts_zero = ["床版"]              # 00になる場合

            # namesから部品名（parts）と数字を抽出
            space = names.find(" ")
            parts = names[:space]  # 部品名
            number = ''.join(extract_number(names))  # 数字
            parts_join = names.replace(number, '') # 符号部分を取得

            # 必要な部分の数字を抽出するロジック
            split_number = ''

            if parts in parts_zero:
                split_number = '00'
            elif len(number) == 4 or int(number[2:]) >= 100:
                if parts in parts_left:
                    split_number = number[:2]
                elif parts in parts_right:
                    split_number = number[2:]
                else:
                    split_number = '00'
            else:
                if parts in parts_left:
                    split_number = number[:3]
                elif parts in parts_right:
                    split_number = number[3:]
                else:
                    split_number = '00'

            result = parts_join + split_number  # 結果を組み立てる
            return result
            # 共通のフィールドを辞書に格納
        infra = Infra.objects.filter(id=pk).first()
        print(f"Infra:{infra}({infra.id})") # 旗揚げチェック(4)
        article = infra.article
        print(f"article:{article}({article.id})") # お試し(2)
        table = Table.objects.filter(infra=infra.id, article=article.id).first()
        print(table) # 旗揚げチェック：お試し（infra/table/dxf/121_2径間番号違い.dxf）
        if not is_multi_list(split_names) and not is_multi_list(damages) and name_length == 1: # 部材名が1つの場合
            for single_damage in damages: 
                parts_name = names[0]
                damage_name = flatten(single_damage)
                #print(f"parts_name1:{parts_name}")
                #print(f"damage_name1:{damage_name}")
                parts_split = process_names(flatten(parts_name))
                data_fields = {
                    'parts_name': parts_name,
                    'damage_name': damage_name,
                    'parts_split': parts_split,
                    'join': join,
                    'picture_number': numbers_only,
                    'this_time_picture': this_time_picture,
                    'last_time_picture': last_time_picture,
                    'textarea_content': textarea_content,
                    'damage_coordinate_x': damage_coordinate_x,
                    'damage_coordinate_y': damage_coordinate_y,
                    'picture_coordinate_x': picture_coordinate_x,
                    'picture_coordinate_y': picture_coordinate_y,
                    'span_number': span_number,
                    'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                    'infra': infra,
                    'article': article,
                    'table': table
                }
                report_data_exists = FullReportData.objects.filter(**data_fields).exists()
                if report_data_exists:
                    print("データが存在しています。")
                else:
                    try:
                        damage_obj, created = FullReportData.objects.update_or_create(**data_fields)
                        damage_obj.save()
                    except IntegrityError:
                        print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")

        elif not is_multi_list(split_names) and not is_multi_list(damages) and name_length >= 2: # 部材名が2つ以上の場合
            if damage_length == 1: # かつ損傷名が1つの場合
                for single_name in split_names:
                    parts_name = single_name
                    damage_name = flatten(damages[0])
                    #print(f"parts_name2:{parts_name}")
                    #print(f"damage_name2:{damage_name}")
                    parts_split = process_names(flatten(parts_name))
                    data_fields = {
                        'parts_name': parts_name,
                        'damage_name': damage_name,
                        'parts_split': parts_split,
                        'join': join,
                        'picture_number': numbers_only,
                        'this_time_picture': this_time_picture,
                        'last_time_picture': last_time_picture,
                        'textarea_content': textarea_content,
                        'damage_coordinate_x': damage_coordinate_x,
                        'damage_coordinate_y': damage_coordinate_y,
                        'picture_coordinate_x': picture_coordinate_x,
                        'picture_coordinate_y': picture_coordinate_y,
                        'span_number': span_number,
                        'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                        'infra': infra,
                        'article': article,
                        'table': table
                    }
                    report_data_exists = FullReportData.objects.filter(**data_fields).exists()
                    if report_data_exists:
                        print("データが存在しています。")
                    else:
                        try:
                            damage_obj, created = FullReportData.objects.update_or_create(**data_fields)
                            damage_obj.save()
                        except IntegrityError:
                            print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")

            elif not is_multi_list(split_names) and not is_multi_list(damages) and damage_length >= 2: # かつ損傷名が2つ以上の場合
                for name in split_names:
                    for damage in damages:
                        parts_name = name
                        damage_name = flatten(damage)
                        #print(f"parts_name3:{parts_name}")
                        #print(f"damage_name3:{damage_name}")
                        parts_split = process_names(flatten(parts_name))
                        data_fields = {
                            'parts_name': parts_name,
                            'damage_name': damage_name,
                            'parts_split': parts_split,
                            'join': join,
                            'picture_number': numbers_only,
                            'this_time_picture': this_time_picture,
                            'last_time_picture': last_time_picture,
                            'textarea_content': textarea_content,
                            'damage_coordinate_x': damage_coordinate_x,
                            'damage_coordinate_y': damage_coordinate_y,
                            'picture_coordinate_x': picture_coordinate_x,
                            'picture_coordinate_y': picture_coordinate_y,
                            'span_number': span_number,
                            'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                            'infra': infra,
                            'article': article,
                            'table': table
                        }
                        report_data_exists = FullReportData.objects.filter(**data_fields).exists()
                        if report_data_exists:
                            print("データが存在しています。")
                        else:
                            try:
                                damage_obj, created = FullReportData.objects.update_or_create(**data_fields)
                                damage_obj.save()
                            except IntegrityError:
                                print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")

        else: # 多重リストの場合
            for i in range(name_length):
                for name in split_names[i]:
                    for damage in damages[i]:
                        parts_name = name
                        damage_name = flatten(damage)
                        #print(f"parts_name4:{parts_name}")
                        #print(f"damage_name4:{damage_name}")
                        parts_split = process_names(flatten(parts_name))
                        data_fields = {
                            'parts_name': parts_name,
                            'damage_name': damage_name,
                            'parts_split': parts_split,
                            'join': join,
                            'picture_number': numbers_only,
                            'this_time_picture': this_time_picture,
                            'last_time_picture': last_time_picture,
                            'textarea_content': textarea_content,
                            'damage_coordinate_x': damage_coordinate_x,
                            'damage_coordinate_y': damage_coordinate_y,
                            'picture_coordinate_x': picture_coordinate_x,
                            'picture_coordinate_y': picture_coordinate_y,
                            'span_number': span_number,
                            'special_links': '/'.join([str(parts_split), str(damage_name), str(span_number)]),
                            'infra': infra,
                            'article': article,
                            'table': table
                        }
                        report_data_exists = FullReportData.objects.filter(**data_fields).exists()
                        if report_data_exists:
                            print("データが存在しています。")
                        else:
                            try:
                                damage_obj, created = FullReportData.objects.update_or_create(**data_fields)
                                damage_obj.save()
                            except IntegrityError:
                                print("ユニーク制約に違反していますが、既存のデータを更新しませんでした。")

    """辞書型の多重リストをデータベースに登録(ここまで)"""
    # path('article/<int:article_pk>/infra/<int:pk>/bridge-table/', views.bridge_table, name='bridge-table')

    #context["damage_table"] = sorted_items
    #return render(request, "infra/bridge_table.html", context)

    # # テンプレートをレンダリング
    # return render(request, 'infra/bridge_table.html', context)
    if "search_title_text" in request.GET:
        # request.GET：検索URL（http://127.0.0.1:8000/article/1/infra/bridge_table/?search_title_text=1径間） 
        search_title_text = request.GET["search_title_text"]
        # 検索URL内のsearch_title_textの値（1径間）を取得する
    else:
        search_title_text = "1径間" # 検索URLにsearch_title_textがない場合
    second_search_title_text = "損傷図"
    
    bridges = FullReportData.objects.filter(infra=pk, span_number=search_title_text)
    print(f"search_title_texts:{search_title_text}")
    print(f"FullReportData:{bridges.first()}")
    # HTMLにまとめて表示するためのグループ化
    grouped_data = []
    for key, group in groupby(bridges, key=attrgetter('join', 'damage_coordinate_x', 'damage_coordinate_y')):
        grouped_data.append(list(group))
        
    photo_grouped_data = []
    for pic_key, pic_group in groupby(bridges, key=attrgetter('this_time_picture', 'span_number')):
        photo_grouped_data.append(list(pic_group))
        
    buttons_count = int(table.infra.径間数) # 数値として扱う
    buttons = list(range(1, buttons_count + 1)) # For loopのためのリストを作成
    # range(一連の整数を作成):range(1からスタート, ストップ引数3 = 2 + 1) → [1, 2](ストップ引数は含まれない)
    print(buttons)
    
    print(f"ボタン:{Table.objects.filter(infra=pk)}")# ボタン:<QuerySet [<Table: Table object (1)>]>
    table_object = Table.objects.filter(infra=pk).first()
    print(f"橋梁番号:{table_object}")# ボタン:Table object (1)
    article_pk = table_object.infra.article.pk
    print(f"案件番号:{article_pk}") # 案件番号:1
    context = {'object': table_object, 'article_pk': article_pk, 'grouped_data': grouped_data, 'photo_grouped_data': photo_grouped_data, 'buttons': buttons}
    # 渡すデータ：　損傷データ　↑　　　       　   joinと損傷座標毎にグループ化したデータ　↑　　　　　　 写真毎にグループ化したデータ　↑ 　　       径間ボタン　↑
    # テンプレートをレンダリング
    return render(request, 'infra/bridge_table.html', context)
    
def ajax_file_send(request, pk):
    if request.method == 'POST': # HTTPリクエストがPOSTメソッドかつ
        if 'upload-file' in request.FILES and 'bridge_id' in request.POST: # アップロードされたファイルがrequest.FILESに入っている場合
            myfile = request.FILES['upload-file'] # 受け取ったファイルをmyfileという変数に代入
            bridge_id = request.POST['bridge_id']
            
            # FileSystemStorageを使ってファイルを保存
            fs = FileSystemStorage() # FileSystemStorageのインスタンスを生成(システム上にファイルを保存する準備)
            filename = fs.save(myfile.name, myfile) # myfileを指定した名前で保存し、保存したファイルの名前をfilename変数に格納
            uploaded_file_url = fs.url(filename) # 保存したファイルにアクセスするためのURLを生成
            
            # データベースを更新する
            try:
                bridge = FullReportData.objects.filter(infra=pk)
                bridge.this_time_picture = uploaded_file_url  # データベースのモデルのフィールドを更新
                bridge.save()
            except FullReportData.DoesNotExist:
                return HttpResponseBadRequest('指定されたブリッジが見つかりませんでした。')

            # success時のレスポンスはJSON形式で返すならこちらを使う
            return JsonResponse({'upload_file_url': uploaded_file_url})
        else:
            # ファイルがPOSTされていない場合はエラーレスポンスを返す
            return HttpResponseBadRequest('添付ファイルが見つかりませんでした。') # File is not attached
    else:
        # POSTメソッドでない場合はエラーレスポンスを返す
        return HttpResponseBadRequest('無効な作業です。') # Invalid request method

# << 写真をアップロードしないでも表示させる >>
def serve_image(request, file_path):
    # セキュリティ: パスを安全にするため制限を加える
    base_dir = os.path.abspath("C:/Users/dobokuka4/Desktop/(件名なし)/案件名/写真")
    # エンコード
    encoded_text = quote(base_dir)
    # URLパスをデコード
    # decoded_path = unquote(file_path)
    # デバッグ用：デコードされたパスの確認
    # print(f"Decoded path: {decoded_path}")
    # OSに依存しない形でのパスの結合
    full_path = os.path.normpath(os.path.join(encoded_text, file_path))
    # デバッグ用：フルパスの確認
    print(f"Full path: {full_path}")
    # ファイルの存在を確認
    if os.path.exists(full_path) and os.path.isfile(full_path):
        return FileResponse(open(full_path, 'rb'))
    else:
        raise Http404("File not found")

# << ファイルアップロード(プライマリーキーで分類分け) >>
def upload_directory_path(instance, filename):
    # プライマリーキーを取得する
    primary_key = instance.pk
    # 'documents/プライマリーキー/filename' のパスを返す
    return 'uploads/{}/{}'.format(primary_key, filename)

# class Upload(models.Model):
#     file = models.FileField(upload_to=upload_directory_path)

# << 所見一覧 >>
def observations_list(request, article_pk, pk):
    context = {}

    table = Table.objects.filter(id=pk).first()
    print(f"table_name:{table}")

    if table.dxf:
        encoded_url_path = table.dxf.url
    decoded_url_path = urllib.parse.unquote(encoded_url_path)
    # 絶対パスと合体
    dxf_filename = os.path.join(settings.BASE_DIR, decoded_url_path.lstrip('/'))
    
    if "search_title_text" in request.GET:
        search_title_text = request.GET["search_title_text"]
    else:
        search_title_text = "1径間"

    second_search_title_text = "損傷図"
    
    sorted_items = create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text)
    """"""
    # 全パーツデータを取得

    infra_name = table.infra.title
    print(f"infra_name:{infra_name}")
    parts_data = PartsNumber.objects.filter(infra=pk)
    print(f"parts_data:{parts_data}")
    
    material_replace_map = {
        "鋼": "S",
        "コンクリート": "C",
        "アスファルト": "A",
        "ゴム": "R",
        "その他": "X",
    }
    
    number_change = {
        '①': '腐食',
        '②': '亀裂',
        '③': 'ゆるみ・脱落',
        '④': '破断',
        '⑤': '防食機能の劣化',
        '⑥': 'ひびわれ',
        '⑦': '剥離・鉄筋露出',
        '⑧': '漏水・遊離石灰',
        '⑨': '抜け落ち',
        '⑩': '補修・補強材の損傷',
        '⑪': '床版ひびわれ',
        '⑫': 'うき',
        '⑬': '遊間の異常',
        '⑭': '路面の凹凸',
        '⑮': '舗装の異常',
        '⑯': '支承部の機能障害',
        '⑰': 'その他',
        '⑱': '定着部の異常',
        '⑲': '変色・劣化',
        '⑳': '漏水・滞水',
        '㉑': '異常な音・振動',
        '㉒': '異常なたわみ',
        '㉓': '変形・欠損',
        '㉔': '土砂詰まり',
        '㉕': '沈下・移動・傾斜',
        '㉖': '洗掘',
    }
    
    lank_order = ['a', 'b', 'c', 'd', 'e']  # ランクの順序をリストで定義
    def get_lank_value(damage_name):
        """damage_nameのランク部分を取得する"""
        if "-" in damage_name:
            return damage_name.split('-')[-1]
        return None
    
    for part in parts_data:
        part_full_name = f"{part.parts_name} {part.symbol}{part.number}"
        span_number = part.span_number + '径間'

        # FullReportDataから該当するデータを取得
        report_data_list = FullReportData.objects.filter(
            parts_name=part_full_name, # FullReportDataのparts_nameオブジェクトがpart_full_name(主桁 Mg0101)と同じ、かつ
            span_number=span_number, # FullReportDataのspan_numberオブジェクトがspan_number(1径間)と同じ、かつ
            infra=part.infra # FullReportDataのinfraオブジェクトがpart.infraと同じ場合
        )  

        for report_data in report_data_list:
            print(f"report_data_list:{report_data.damage_name}")
            print(f"picture:{report_data.this_time_picture}")

            damage_list_material = "" # 空のdamage_list_materialを用意
            for m in part.material.all(): # part.materialの全データを取得し「m」変数に入れる
                damage_list_material += m.材料 + "," # 「m」の材料フィールドを指定してdamage_list_materialに入れる
                
            elements = damage_list_material.split(',')
            replaced_elements = [material_replace_map.get(element, element) for element in elements] # それぞれの要素を置換辞書に基づいて変換します
            damage_list_materials = ','.join(replaced_elements) # カンマで結合します。
            
            damage_name = report_data.damage_name.split('-')[0] if '-' in report_data.damage_name else report_data.damage_name
            if damage_name == "NON":
                damage_name = damage_name
            elif damage_name[0] != '⑰':
                damage_name = number_change[damage_name[0]]
            else:
                damage_name = damage_name[1:] # ⑦剥離・鉄筋露出から先頭の一文字を省く
                
            damage_lank = report_data.damage_name.split('-')[1] if '-' in report_data.damage_name else report_data.damage_name
            
            # DamageListに必要なフィールドを含むインスタンスを作成
            # << 損傷一覧(Excel)用データ登録 >>
            damage_list_entry = DamageList(
                parts_name = part.parts_name, # 主桁
                symbol = part.symbol, # Mg
                number = part.number, # 0101
                material = damage_list_materials[:-1], # 最後のコンマが不要なため[-1:]（S,C）
                main_parts = "〇" if part.main_frame else "", # 主要部材のフラグ
                damage_name = damage_name, # 剥離・鉄筋露出
                damage_lank = damage_lank, # d
                span_number = part.span_number,
                infra = part.infra
            )

            try:
                # DamageListインスタンスを保存
                damage_list_entry.save()
                
            except IntegrityError:
                # 重複データがある場合の処理
                print("データが存在しています。")
                # 必要に応じてログを記録したり、他の処理を追加したりできます
                # continue  # 次のループに進む
            
    """所見用のクラス登録"""
    damage_comments = defaultdict(lambda: {'damage_lanks': [], 'this_time_pictures': []})

    for part in parts_data:
        part_full_name = f"{part.parts_name} {part.symbol}{part.number}"
        span_number = part.span_number + '径間'

        report_data_list = FullReportData.objects.filter(
            parts_name=part_full_name,
            span_number=span_number,
            infra=part.infra
        )

        for report_data in report_data_list:
            main_parts_list_left = ["主桁", "PC定着部"]
            main_parts_list_right = ["横桁", "橋台"]
            main_parts_list_zero = ["床版"]

            parts_name = f"{part.parts_name} {part.number}"

            if any(word in parts_name for word in main_parts_list_left):
                left = parts_name.find(" ")
                number2 = parts_name[left+1:]
                number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                result_parts_name = parts_name[:left] + " " + number_part[:2]
            elif any(word in parts_name for word in main_parts_list_right):
                right = parts_name.find(" ")
                number2 = parts_name[right+1:]
                number_part = re.search(r'[A-Za-z]*(\d+)', number2).group(1)
                result_parts_name = parts_name[:right] + " " + number_part[2:] if len(number_part) < 5 else number_part[2:]
            elif any(word in parts_name for word in main_parts_list_zero):
                right = parts_name.find(" ")
                result_parts_name = parts_name[:right] + " 00"
            else:
                right = parts_name.find(" ")
                result_parts_name = parts_name[:right]

            damage_name = report_data.damage_name.split('-')[0] if '-' in report_data.damage_name else report_data.damage_name
            if damage_name == "NON":
                damage_name = damage_name
            elif damage_name[0] != '⑰':
                damage_name = number_change[damage_name[0]]
            else:
                damage_name = damage_name[1:] 
            damage_lank = report_data.damage_name.split('-')[1] if '-' in report_data.damage_name else report_data.damage_name    

            damage_comments[(result_parts_name, damage_name)]['damage_lanks'].append(damage_lank)
            damage_comments[(result_parts_name, damage_name)]['this_time_pictures'].append(report_data.this_time_picture)
            
            damage_comment_material = ""
            for m in part.material.all():
                damage_comment_material += m.材料 + ","
            elements = damage_comment_material.split(',')
            replaced_elements = [material_replace_map.get(element, element) for element in elements]
            damage_comment_materials = ','.join(replaced_elements)

            damage_comments[(result_parts_name, damage_name)]['material'] = damage_comment_materials[:-1]
            damage_comments[(result_parts_name, damage_name)]['main_parts'] = "〇" if part.main_frame else ""
            damage_comments[(result_parts_name, damage_name)]['span_number'] = part.span_number
            damage_comments[(result_parts_name, damage_name)]['infra'] = part.infra

    for (result_parts_name, damage_name), data in damage_comments.items():
        damage_lanks = data['damage_lanks']
        damage_max_lank = max(damage_lanks)
        damage_min_lank = min(damage_lanks)
        
        combined_pictures = ','.join(str(picture) for picture in set(data['this_time_pictures']) if picture is not None) # 重複なし

        try:
            damage_comment_entry = DamageComment(
                parts_name=result_parts_name,
                material=data['material'],
                main_parts=data['main_parts'],
                damage_name=damage_name,
                damage_max_lank=damage_max_lank,
                damage_min_lank=damage_min_lank,
                this_time_picture=combined_pictures,
                span_number=data['span_number'],
                infra=data['infra']
            )
            damage_comment_entry.save()

        except IntegrityError:
            # 重複データがある場合の処理
            print("データが存在しています。")
            # 必要に応じてログを記録したり、他の処理を追加したりできます
            continue  # 次のループに進む
        
    # span_numberの順かつ、replace_nameの順かつ、parts_numberの順かつ、numberの順に並び替え 
    sorted_data = DamageComment.objects.filter(infra=pk).order_by('span_number', 'replace_name', 'parts_number', 'number')
    """ここまで"""
    
    if "search_title_text" in request.GET:
        search_title_text = request.GET["search_title_text"]

    else:
        search_title_text = "1径間"
    
    span_create_number = search_title_text.replace("径間", "")
    print(span_create_number)
    filtered_bridges = DamageComment.objects.filter(infra=pk, span_number=span_create_number).order_by('span_number', 'replace_name', 'parts_number', 'number')
    print(f"bridges:{filtered_bridges}")
    buttons_count = int(table.infra.径間数) # 数値として扱う
    buttons = list(range(1, buttons_count + 1)) # For loopのためのリストを作成
    # range(一連の整数を作成):range(1からスタート, ストップ引数3 = 2 + 1) → [1, 2](ストップ引数は含まれない)
    print(buttons)
    
    print(f"所見ボタン:{DamageComment.objects.filter(infra=pk)}")# ボタン:<QuerySet [<Table: Table object (15)>]>
    print(f"所見ボタン:{DamageComment.objects.filter(infra=pk).first()}")# ボタン:Table object (18)(QuerySetのままだとうまく動作しない)
    #   1(径間)  ,      1(主桁)  ,        01     ,    6(ひびわれ)

    observer_object = Table.objects.filter(id=pk).first()
    print(f"橋梁番号:{observer_object}") # 橋梁番号:Table object (1)
    article_pk = observer_object.article.pk
    print(f"案件番号:{article_pk}") # 案件番号:1
    return render(request, 'observer_list.html', {'object': observer_object, 'article_pk': article_pk, 'data': filtered_bridges, 'article_pk': article_pk, 'pk': pk, 'buttons': buttons})

# << 所見コメントの登録 >>
def damage_comment_edit(request, pk):
    if request.method == "POST":
        # TODO: 編集を受け付ける
        # DamageComment の idを受け取る。
        # URL：path('damage_comment_edit/<int:pk>/', views.damage_comment_edit , name="damage_comment_edit")
        damage_comment = DamageComment.objects.get(id=pk) # idが同じDamageCommentデータを取得(int:pk 1種類のidが必要)
        print(damage_comment)
        form = DamageCommentEditForm(request.POST, instance=damage_comment)
     # ユーザーが送信したPOSTデータをFormに渡す ↑　　　　　　　　↑ 編集するオブジェクト
        print("編集します。")

        if form.is_valid(): # バリデーション
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)
        # リダイレクト処理  　　　　　　　　　　　　　↓　damage_commentクラス → infraフィールド(infraクラスに移る) → articleフィールド(articleクラスに移る)からarticle.idを取得
        return redirect("observations-list", damage_comment.infra.article.id, damage_comment.infra.id )
        # damage_commentクラス → infraフィールド(infraクラスに移る)からinfra.idを取得　↑　　　　　(int:article_pk、int:pk 2種類のidが必要)
        # path('article/<int:article_pk>/infra/<int:pk>/observations/', views.observations_list, name='observations-list')
        
# << Ajaxを使用した所見のリアルタイム保存 >>
@csrf_protect  # CSRF保護を有効にする
def save_comment(request, pk):
    if request.method == "POST":
        damage_comment = get_object_or_404(DamageComment, id=pk)
        form = DamageCommentEditForm(request.POST, instance=damage_comment)
        
        if form.is_valid():
            form.save()
            return JsonResponse({"status": "success"})
        return JsonResponse({"status": "error", "errors": form.errors})
    return JsonResponse({"status": "error", "message": "Invalid request method."})

# << Ajaxを使用した損傷写真帳のリアルタイム保存 >>
@csrf_protect  # CSRF保護を有効にする
def update_full_report_data(request, pk):
    if request.method == 'POST':
        full_report_data = get_object_or_404(FullReportData, id=pk)
        
        fields_to_update = ['measurement', 'damage_size', 'classification', 'pattern']
        data_updated = False
        
        for field in fields_to_update:
            if field in request.POST:
                setattr(full_report_data, field, request.POST[field])
                data_updated = True

        if data_updated:
            full_report_data.save()
            return JsonResponse({'status': 'success'})
        else:
            return JsonResponse({'status': 'error', 'errors': 'No valid data to update'})
    return JsonResponse({'status': 'invalid request'})

# << 対策区分のボタンを保存 >>
def damage_comment_jadgement_edit(request, pk):
    if request.method == "POST":
        #TODO: 編集を受け付ける。
        # DamageComment の idを受け取る。
        print("DamageCommentJadgementEditForm 発動。")
        damage_comment = DamageComment.objects.get(id=pk)
        form = DamageCommentJadgementEditForm(request.POST, instance=damage_comment)
        
        if form.is_valid():
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)

        return redirect("observations-list", damage_comment.infra.article.id, damage_comment.infra.id )

# << 損傷原因のボタンを保存 >>
def damage_comment_cause_edit(request, pk):
    if request.method == "POST":
        #TODO: 編集を受け付ける。
        # DamageComment の idを受け取る。
        print("DamageCommentCauseEditForm 発動。")
        damage_comment_cause = DamageComment.objects.get(id=pk)
        form = DamageCommentCauseEditForm(request.POST, instance=damage_comment_cause)
        
        if form.is_valid():
            form.save()
            print("編集保存完了")
        else:
            print(form.errors)

        return redirect("observations-list", damage_comment_cause.infra.article.id, damage_comment_cause.infra.id )

# << 名前とアルファベットを紐付け >>
def names_list(request, article_pk):
    
    alphabet_list = request.POST.getlist("name_alphabet")
    #alphabet_dict = {alphabet_list[i]: alphabet_list[i+1] for i in range(0, len(alphabet_list), 2)}
    #print(alphabet_dict)
    for i in range(0, len(alphabet_list), 2):
        dic = {}
        dic["name"] = alphabet_list[i]
        dic["alphabet"] = alphabet_list[i+1]
        dic["article"] = article_pk
        
        form = NameEntryForm(dic)

        if form.is_valid():
            form.save()
        else:
            print(form.errors) # 入力フォームのエラー内容を表示
            
    # NameEntryには、article_id のフィールドはない、 article に変更
    # name_entries = NameEntry.objects.filter(article_id=article_pk)
    name_entries = NameEntry.objects.filter(article=article_pk)
    
    return render(request, 'names_list.html', {'article_pk': article_pk, "form": NameEntryForm(), 'name_entries': name_entries})

# << 登録した名前を削除 >>
def delete_name_entry(request, entry_id):
    entry = get_object_or_404(NameEntry, pk=entry_id)
    article_pk = entry.article.pk  # 事前に記事のPKを取得
    if request.method == 'POST':    
        entry.delete()
    name_entries = NameEntry.objects.filter(article=article_pk)
    return render(request, 'names_list.html', {'article_pk': article_pk, "form": NameEntryForm(), 'name_entries': name_entries})

# << 番号登録 >>
def number_list(request, article_pk, pk):

    # 同じname属性の値をすべて取り出す
    serial_numbers = request.POST.getlist("serial_number") # ['0101', '0103', '0201', '0203']
    single_numbers = request.POST.getlist("single_number") # ['0101', '0201', '0301', '0401']
    
    new_serial_numbers = []
    #    初期値を0 ↓     回数分 ↓           ↓ 2ずつ足す(0101(index:0),0201(index:2))
    for i in range(0, len(serial_numbers), 2):
        new_serial_numbers.append(serial_numbers[i] + "~" + serial_numbers[i+1])
        #                          0101(index:0) ↑          0103(index:1+1) ↑
        #                          0201(index:2) ↑          0203(index:2+1) ↑
    print(new_serial_numbers) # ['0101~0103', '0201~0203']
    
    # 単一の番号、連続の番号 を部材名と紐付けて保存
    for new_serial_number in new_serial_numbers:
        print(new_serial_number)
        if "~" in new_serial_number and len(new_serial_number) >= 5: # 01～02(5文字)
            # new_serial_number = "0101~0205"
            one = new_serial_number.find("~")

            start_number = new_serial_number[:one]
            end_number = new_serial_number[one+1:]

            # 最初の2桁と最後の2桁を取得
            start_prefix = start_number[:2]
            start_suffix = start_number[2:]
            end_prefix = end_number[:2]
            end_suffix = end_number[2:]

            first_elements = []
            # 決められた範囲内の番号を一つずつ追加
            for prefix in range(int(start_prefix), int(end_prefix)+1):
                for suffix in range(int(start_suffix), int(end_suffix)+1):
                    number_items = "{:02d}{:02d}".format(prefix, suffix)
                    dic = {} # forms.pyにも入れないと自動登録ができない
                    dic["number"] = number_items
                    dic["parts_name"] = request.POST.get("parts_name")
                    dic["symbol"] = request.POST.get("symbol")
                    dic["material"] = request.POST.getlist("material")
                    dic["span_number"] = request.POST.get("span_number")
                    dic["main_frame"] = request.POST.get("main_frame") == 'on'
                    dic["infra"] = pk # infraとの紐付け
                    dic["article"] = article_pk
                    print(f"new_serial_number:{number_items}")
                    
                    # 1個ずつバリデーションして保存する
                    form = PartsNumberForm(dic)

                    if form.is_valid():
                        form.save()
                        parts_number = form.save()
                        parts_number.material.set(request.POST.getlist("material"))
                    else:
                        print(form.errors) # 入力フォームのエラー内容を表示
                        
    for single_number in single_numbers:
        if single_number.isdigit():
            dic = {}
            dic["number"] = single_number
            dic["parts_name"] = request.POST.get("parts_name")
            dic["symbol"] = request.POST.get("symbol")
            dic["material"] = request.POST.getlist("material")
            dic["span_number"] = request.POST.get("span_number")
            dic["main_frame"] = request.POST.get("main_frame") == 'on'
            dic["infra"] = pk # infraとの紐付け
            dic["article"] = article_pk 
            print(single_number)

            # 1個ずつバリデーションして保存する
            form    = PartsNumberForm(dic)

            if form.is_valid():
                form.save()
                parts_number = form.save()
                parts_number.material.set(request.POST.getlist("material"))
            else:
                print(form.errors)
        
    parts_names = PartsName.objects.all()
    create_number_list = PartsNumber.objects.filter(infra=pk)
    
    grouped_parts = defaultdict(list)
    for accordion_list in create_number_list:
        title = f"{accordion_list.parts_name.部材名}（{accordion_list.symbol}）{accordion_list.get_material_list()} {accordion_list.span_number}径間"
        grouped_parts[title].append(accordion_list.number)

    create_number_list = PartsNumber.objects.filter(infra=pk).first()
    print(f"橋梁番号:{create_number_list}") # 橋梁番号:Table object (1)
    article_pk = create_number_list.article.pk
    print(f"案件番号:{article_pk}") # 案件番号:1
    return render(request, 'number_entry.html', {'object': create_number_list, 'article_pk': article_pk, 'pk': pk, "form": PartsNumberForm(), "parts_names":parts_names, 'create_number_list': create_number_list, 'grouped_parts': grouped_parts.items()})
    # return render(request, 'observer_list.html', {'object': observer_object, 'article_pk': article_pk, 'data': filtered_bridges, 'article_pk': article_pk, 'pk': pk, 'buttons': buttons})

# << 登録した番号を削除 >>
def delete_number_entry(request, article_entry_id, entry_id):
    entry = get_object_or_404(PartsNumber, article_pk=article_entry_id, pk=entry_id) # pkに対応するPartsNameオブジェクトをデータベースから取得
    print(entry)
    article_pk = entry.article.pk # entryのarticleのpkを取得（削除前に行う）
    # pk = entry.pk
    if request.method == 'POST':
        entry.delete() # データベースから削除
    parts_names = PartsNumber.objects.filter(article=article_pk)
    return render(request, 'number_entry.html', {'article_pk': article_pk, "form": PartsNumberForm(), "parts_names":parts_names})

# 部材名と記号を紐付けるAjaxリクエスト
def get_symbol(request):
    part_id = request.GET.get('part_id')
    try:
        parts_name = PartsName.objects.get(id=part_id)
        return JsonResponse({'symbol': parts_name.記号})
    except PartsName.DoesNotExist:
        return JsonResponse({'error': 'PartsName not found'}, status=404)
    
# 番号表示  
def number_view(request):
    # PartsNumberモデルから1件データを取り出し
    parts_number = PartsNumber.objects.get(id=1)
    # 抽出した数字を文字列として結合
    result = ""
    # 4桁 か 4桁~4桁 のいずれか
    if len(parts_number.number) == 4:
        # 4桁
        result  = parts_number.number
    else:
        # 4桁~4桁
        # ~ で区切る必要がある。 [ "3000","3000" ]
        numbers = parts_number.number.split("~")
        start   = numbers[0]
        end     = numbers[1]
        # 最初の2桁と最後の2桁を取得
        start_prefix = start[:2]
        start_suffix = start[2:]
        end_prefix = end[:2]
        end_suffix = end[2:]

        for prefix in range(int(start_prefix), int(end_prefix)+1):
            for suffix in range(int(start_suffix), int(end_suffix)+1):
                result += "{:02d}{:02d}\n".format(prefix, suffix)

    print(result)

# << 橋梁緒言の選択肢 >>
def infraregulations_view(request):
    form = BridgeCreateForm()
    regulations = Regulation.objects.all()
    context = {
        'form': form,
        'regulations': regulations,
    }
    return render(request, 'infra_create.html', context)

def infraloadWeights_view(request):
    form = BridgeCreateForm()
    loadWeights = LoadWeight.objects.all()
    context = {
        'form': form,
        'loadWeights': loadWeights,
    }
    return render(request, 'infra_create.html', context)

def infraloadGrades_view(request):
    form = BridgeCreateForm()
    loadGrades = LoadGrade.objects.all()
    context = {
        'form': form,
        'loadGrades': loadGrades,
    }
    return render(request, 'infra_create.html', context)

def infrarulebooks_view(request):
    form = BridgeCreateForm()
    rulebooks = Rulebook.objects.all()
    context = {
        'form': form,
        'rulebooks': rulebooks,
    }
    return render(request, 'infra_create.html', context)

def infraapproachs_view(request):
    form = BridgeCreateForm()
    approachs = Approach.objects.all()
    context = {
        'form': form,
        'approachs': approachs,
    }
    return render(request, 'infra_create.html', context)

def infrathirdpartys_view(request):
    form = BridgeCreateForm()
    thirdpartys = Thirdparty.objects.all()
    context = {
        'form': form,
        'thirdpartys': thirdpartys,
    }
    return render(request, 'infra_create.html', context)

def infraunderConditions_view(request):
    form = BridgeCreateForm()
    underConditions = UnderCondition.objects.all()
    context = {
        'form': form,
        'underConditions': underConditions,
    }
    return render(request, 'infra_create.html', context)

def infraregulations_view(request):
    form = BridgeUpdateForm()
    regulations = Regulation.objects.all()
    context = {
        'form': form,
        'regulations': regulations,
    }
    return render(request, 'infra_update.html', context)

def infraloadWeights_view(request):
    form = BridgeUpdateForm()
    loadWeights = LoadWeight.objects.all()
    context = {
        'form': form,
        'loadWeights': loadWeights,
    }
    return render(request, 'infra_update.html', context)

def infraloadGrades_view(request):
    form = BridgeUpdateForm()
    loadGrades = LoadGrade.objects.all()
    context = {
        'form': form,
        'loadGrades': loadGrades,
    }
    return render(request, 'infra_update.html', context)

def infrarulebooks_view(request):
    form = BridgeUpdateForm()
    rulebooks = Rulebook.objects.all()
    context = {
        'form': form,
        'rulebooks': rulebooks,
    }
    return render(request, 'infra_update.html', context)

def infraapproachs_view(request):
    form = BridgeUpdateForm()
    approachs = Approach.objects.all()
    context = {
        'form': form,
        'approachs': approachs,
    }
    return render(request, 'infra_update.html', context)

def infrathirdpartys_view(request):
    form = BridgeUpdateForm()
    thirdpartys = Thirdparty.objects.all()
    context = {
        'form': form,
        'thirdpartys': thirdpartys,
    }
    return render(request, 'infra_update.html', context)

def infraunderConditions_view(request):
    form = BridgeUpdateForm()
    underConditions = UnderCondition.objects.all()
    context = {
        'form': form,
        'underConditions': underConditions,
    }
    return render(request, 'infra_update.html', context)

parts_name_priority_list = ['主桁', '横桁', '床版']

# カスタムソートキー関数
def custom_sort_key(record):
    # parts_nameリストの優先度を求めるためのインデックス
    parts_name_priority = next((i for i, part in enumerate(parts_name_priority_list) if part in record.parts_name), len(parts_name_priority_list))
    return (int(record.span_number), parts_name_priority)

# << 指定したInfra(pk)に紐づくTableのエクセルの出力 >>
def excel_output(request, article_pk, pk):
    
    # 元のファイルのパス（例: `base.xlsm`）
    original_file_path = 'base.xlsm'
    # エクセルファイルを読み込む
    wb = openpyxl.load_workbook(original_file_path, keep_vba=True)
    
    # << Django管理サイトからデータを取得（その１用） >>
    no01_records = Infra.objects.filter(id=pk)
    ws = wb['その１']
    for record in no01_records:
        print(record.title)
        ws[f'F6'] = record.title # 〇〇橋
        ws[f'O10'] = record.橋長
        ws[f'BC5'] = record.橋梁コード
        ws[f'BF11'] = record.交通量
        ws[f'BF13'] = record.大型車混入率
        # 活荷重を処理
        load_weights = ', '.join([str(weight) for weight in record.活荷重.all()])
        ws['X10'] = load_weights
        # 等級を処理
        load_grades = ', '.join([str(grade) for grade in record.等級.all()])
        ws['AD10'] = load_grades
        # 適用示方書を処理
        rulebooks = ', '.join([str(rulebook) for rulebook in record.適用示方書.all()])
        ws['AK10'] = rulebooks

    # << Django管理サイトからデータを取得（その７、８用） >>
    no0708_records = DamageComment.objects.filter(infra=pk)
    # 並び替えて出力
    sorted_records = sorted(no0708_records, key=custom_sort_key)
    # カウンタ変数をシートごとに用意
    span = 1
    i07, i08 = 0, 0
    initial_row07, initial_row08= 13, 13 # 1つ目の入力位置
    
    for record in sorted_records:
        print(f"出力レコード:{record}")
        print(f"　径間:{span}")
        if int(record.span_number) == span + 1:
            span = int(record.span_number)
            initial_row07 = initial_row07 + 8 * math.ceil(i07 / 8) # 13+8×(ページ数)
            initial_row08 = initial_row08 + 8 * math.ceil(i08 / 8) # 13,21,29(+8)
            i07, i08 = 0, 0
        if int(record.span_number) == span:
            if record.main_parts == "〇":
                ws = wb['その７']
                row = initial_row07 + i07 # 行は13から
                i07 += 1
            else:
                ws = wb['その８']
                row = initial_row08 + i08 # 行は13から
                i08 += 1
            print(f"　エクセル出力:{record.comment_parts_name}{record.parts_number}{record.damage_name}{record.jadgement}")
            ws[f'F{row}'] = record.comment_parts_name # 主桁
            ws[f'J{row}'] = record.parts_number # 01
            ws[f'D{row}'] = record.material # S,C
            ws[f'L{row}'] = record.damage_max_lank # e
            ws[f'N{row}'] = record.damage_min_lank # b
            ws[f'BO{row}'] = record.span_number # 径間番号
            
            if record.damage_name != "NON":
                if record.jadgement == "C1":
                    jadgement_position = f'S{row}'
                elif record.jadgement == "C2":
                    jadgement_position = f'V{row}'
                elif record.jadgement == "M":
                    jadgement_position = f'Z{row}'
                elif record.jadgement == "E1":
                    jadgement_position = f'AD{row}'
                elif record.jadgement == "E2":
                    jadgement_position = f'AH{row}'
                elif record.jadgement == "S1":
                    jadgement_position = f'AK{row}'
                elif record.jadgement == "S2":
                    jadgement_position = f'AN{row}'
                else:
                    jadgement_position = f'P{row}'
            
                ws[jadgement_position] = record.damage_name # 腐食
                ws[f'AU{row}'] = record.cause # 損傷原因「経年変化」  
            
                if record.comment != "":
                    choice_comment = record.comment
                else:
                    choice_comment = record.auto_comment
                
                ws[f'BC{row}'] = choice_comment # 〇〇が見られる。
            else:
                ws[f'BC{row}'] = "健全である。"

    # << （その１０） >>
    no10_records = FullReportData.objects.filter(infra=pk)
    ws = wb['その１０']
    print(f"最大径間数：{span}")
    print(f"最大径間数：{int(span)}")
    # << セル位置を作成 >>
    picture_and_spannumber_row = 9 # 部材名・要素番号
    partsname_and_number_row = 10 # 部材名・要素番号
    damagename_and_lank_row = 11 # 損傷の種類・損傷程度
    picture_start_row = 13 # 損傷写真
    lasttime_lank_row = 15 # 前回損傷程度
    damage_memo_row = 17 # 損傷メモ
    step = 14
    output_data = len(no10_records)
    num_positions = math.ceil(output_data/3) + int(span) * 2 # 横3列で割って何行になるか
    
    # 関連する列を定義
    picture_columns = ["E", "AE", "BE"] # 写真列
    left_columns = ["I", "AI", "BI"] # 左列
    right_columns = ["R", "AR", "BR"] # 右列
    bottom_columns = ["T", "AT", "BT"] # 前回程度+メモ
    # セル位置のリストを生成
    join_spannumber_cell = [f"{col}{picture_and_spannumber_row + i * step}" for i in range(num_positions) for col in right_columns] # 径間番号
    join_partsname_cell = [f"{col}{partsname_and_number_row + i * step}"    for i in range(num_positions) for col in left_columns] # 部材名
    join_number_cell = [f"{col}{partsname_and_number_row + i * step}"       for i in range(num_positions) for col in right_columns] # 要素番号
    join_damagename_cell = [f"{col}{damagename_and_lank_row + i * step}"    for i in range(num_positions) for col in left_columns] # 損傷の種類
    join_lank_cell = [f"{col}{damagename_and_lank_row + i * step}"          for i in range(num_positions) for col in right_columns] # 損損傷程度
    join_picture_cell = [f"{col}{picture_start_row + i * step}"             for i in range(num_positions) for col in picture_columns] # 損傷写真
   #join_lasttime_lank_cell = [f"{col}{lasttime_lank_row + i * step}"       for i in range(num_positions) for col in bottom_columns] # 前回損傷程度
    join_damage_memo_cell = [f"{col}{damage_memo_row + i * step}"           for i in range(num_positions) for col in bottom_columns] # 損傷メモ

    span = 1
    page_count = 1
    i10 = 0
    initial_row10 = 9 # 1つ目の入力位置
    
    def hide_sheet_copy_and_paste(wb, sheet_name):
        """シートを再表示してコピーその後非表示に設定"""

        hide_sheet = wb['ページ１０']
        hide_sheet.sheet_state = 'visible'

        # コピーする行の範囲を指定します
        copy_start_row = 2
        copy_end_row = 29

        # コピーする行のデータとスタイルを保持するリストを作成します
        rows_to_copy = []
        merges_to_keep = []

        for row_idx in range(copy_start_row, copy_end_row + 1):
            row_data = []
            for cell in hide_sheet[row_idx]:
                cell_data = {
                    'value': cell.value,
                    'font': copy(cell.font),
                    'border': copy(cell.border),
                    'fill': copy(cell.fill),
                    'number_format': cell.number_format,
                    'protection': copy(cell.protection),
                    'alignment': copy(cell.alignment)
                }
                row_data.append(cell_data)
            row_data.append(hide_sheet.row_dimensions[row_idx].height)
            rows_to_copy.append(row_data)

        # 元のシートのセル結合情報を取得
        for merge in hide_sheet.merged_cells.ranges:
            if (copy_start_row <= merge.min_row <= copy_end_row) or \
                (copy_start_row <= merge.max_row <= copy_end_row):
                merges_to_keep.append(copy(merge))
        
        sheet = ws
        
        # コピー先の行を挿入します
        # A列の一番下の行番号を取得
        max_row = sheet.max_row
        while sheet['A' + str(max_row)].value is None and max_row > 0:
            max_row -= 1
        insert_at_row = max_row
        # print(f"max_row：{max_row}")
        
        # シフトする行の高さを保持するリストを作成します
        heights = []
        for row_idx in range(insert_at_row, sheet.max_row + 1):
            heights.append(sheet.row_dimensions[row_idx].height)
        
        # 指定行から下の行をシフト
        sheet.insert_rows(insert_at_row, amount=(copy_end_row - copy_start_row + 1))

        # 行の高さを元に戻す
        for i, height in enumerate(heights):
            sheet.row_dimensions[insert_at_row + i + (copy_end_row - copy_start_row + 1)].height = height
        
        # コピーされた行を挿入
        for i, row_data in enumerate(rows_to_copy):
            new_row = insert_at_row + i
            for j, cell_data in enumerate(row_data[:-1]):
                cell = sheet.cell(row=new_row, column=j + 1)
                cell.value = cell_data['value']
                cell.font = cell_data['font']
                cell.border = cell_data['border']
                cell.fill = cell_data['fill']
                cell.number_format = cell_data['number_format']
                cell.protection = cell_data['protection']
                cell.alignment = cell_data['alignment']
            sheet.row_dimensions[new_row].height = row_data[-1]

        # セル結合をコピー
        for merged_range in merges_to_keep:
            new_min_row = merged_range.min_row - copy_start_row + insert_at_row
            new_max_row = merged_range.max_row - copy_start_row + insert_at_row
            new_merge_range = "{}{}:{}{}".format(
                openpyxl.utils.get_column_letter(merged_range.min_col),
                new_min_row,
                openpyxl.utils.get_column_letter(merged_range.max_col),
                new_max_row
            )
            sheet.merge_cells(new_merge_range)
            
        # 最大行を取得
        max_row = sheet.max_row
        # 印刷範囲の設定を修正
        start_col = "A"
        end_col = 'CD'
        print_area = f"{start_col}1:{end_col}{max_row}"
        sheet.print_area = print_area    
        
        hide_sheet.sheet_state = 'hidden'

    for record in no10_records:
        span_number = record.span_number.replace('径間', '')
        print(f"その10レコード数：{len(no10_records)}")
        if int(span_number) == span + 1:
            span = int(span_number)
            initial_row10 = initial_row10 + 28 * math.ceil(i10 / 6) # 9+28×(ページ数)
            ws[join_spannumber_cell[i10]] = span
            
        if int(span_number) == span and record.this_time_picture != "": # 同じ径間で写真データが含まれている場合
            if i10 % 6 == 5 and i10 > 10:
                hide_sheet_copy_and_paste(wb, ws) # プログラム4｜1ページ増やすマクロを実行

            print(f"対象数:{i10}/{len(no10_records)}")
            if record.parts_name in " ":
                # 文字列を空白で分ける
                split_parts = record.parts_name.split()
                # 最初の部分「アクション-1」を取得
                parts_name = split_parts[0]
                # 2番目の部分から数値部分だけを取得
                parts_number = re.search(r'\d+', split_parts[1]).group()
            else:
                parts_name = ""
                parts_number = ""
            if record.damage_name in "-":
                # 置き換え用の辞書
                number_change = {
                    '①': '腐食',
                    '②': '亀裂',
                    '③': 'ゆるみ・脱落',
                    '④': '破断',
                    '⑤': '防食機能の劣化',
                    '⑥': 'ひびわれ',
                    '⑦': '剥離・鉄筋露出',
                    '⑧': '漏水・遊離石灰',
                    '⑨': '抜け落ち',
                    '⑩': '補修・補強材の損傷',
                    '⑪': '床版ひびわれ',
                    '⑫': 'うき',
                    '⑬': '遊間の異常',
                    '⑭': '路面の凹凸',
                    '⑮': '舗装の異常',
                    '⑯': '支承部の機能障害',
                    '⑰': 'その他',
                    '⑱': '定着部の異常',
                    '⑲': '変色・劣化',
                    '⑳': '漏水・滞水',
                    '㉑': '異常な音・振動',
                    '㉒': '異常なたわみ',
                    '㉓': '変形・欠損',
                    '㉔': '土砂詰まり',
                    '㉕': '沈下・移動・傾斜',
                    '㉖': '洗掘',
                }

                first_char = record.damage_name[0] # 先頭の1文字を取得                
                
                damage_name = number_change.get(first_char, "") # 辞書で値を取得
                damage_lank = record.damage_name[-1]
            else:
                damage_name = ""
                damage_lank = ""
            ws[join_partsname_cell[i10]] = parts_name # 部材名
            ws[join_number_cell[i10]] = parts_number # 要素番号
            ws[join_damagename_cell[i10]] = damage_name # 損傷の種類
            ws[join_lank_cell[i10]] = damage_lank # 損傷の程度
            ws[join_picture_cell[i10]] = record.this_time_picture # 損傷写真
            ws[join_damage_memo_cell[i10]] = record.textarea_content # メモ
            i10 += 1

    # << Django管理サイトからデータを取得（その１１、１２用） >>
    no1112_records = DamageList.objects.filter(infra=pk)
    # 並び替え
    sorted_records = sorted(no1112_records, key=custom_sort_key)
    span = 1
    i11, i12 = 0, 0
    initial_row11, initial_row12 = 10, 10 # 1つ目の入力位置

    for record in sorted_records:
        print(f"出力レコード:{record}")
        print(f"　径間:{span}")
        if int(record.span_number) == span + 1:
            span = int(record.span_number)
            initial_row11 = initial_row11 + 18 * math.ceil(i11 / 18) # 10+18×(ページ数)
            initial_row12 = initial_row12 + 18 * math.ceil(i12 / 18) # 10,28,46(+18)
            i11, i12 = 0, 0
        if int(record.span_number) == span:
            if record.main_parts == "〇":
                ws = wb['その１１']
                row = initial_row11 + i11 # 行は10から
                i11 += 1
            else:
                ws = wb['その１２']
                row = initial_row12 + i12 # 行は10から
                i12 += 1
            print(f"　エクセル出力:{record.parts_name}{record.damage_name}{record.span_number}")
            ws[f'H{row}'] = record.parts_name # 主桁
            ws[f'T{row}'] = record.number # 0101
            ws[f'E{row}'] = record.material # S,C
            ws[f'AR{row}'] = record.damage_name # 腐食
            ws[f'X{row}'] = record.damage_lank # d
            ws[f'BE{row}'] = record.classification # 分類「1」
            ws[f'AO{row}'] = record.pattern # パターン「6」
            ws[f'BL{row}'] = record.span_number # 径間番号
    
    # 現在の日時を取得してファイル名に追加
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    # 新しいファイル名の生成
    new_filename = f"Macro{timestamp}_{original_file_path}"
    # デスクトップのパス
    # desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    # 保存するファイルのフルパス
    # save_path = os.path.join(desktop_path, new_filename)

    #メモリ空間内に保存
    virtual = BytesIO()
    wb.save(virtual)
    #バイト文字列からバイナリを作る
    binary = BytesIO(virtual.getvalue())
    return FileResponse(binary, filename = new_filename)

# << 指定したInfra(pk)に紐づくTableのエクセルの出力 >>
def dxf_output(request, article_pk, pk):
    # pythoncom.CoInitialize() # COMライブラリを初期化
    #try:
        # 指定したInfraに紐づく Tableを取り出す
        table = Table.objects.filter(infra=pk).first()
        # print(f"テーブル：{table.dxf.url}") # 相対パス
        
        # 絶対パスに変換
        encoded_url_path = table.dxf.url
        decoded_url_path = urllib.parse.unquote(encoded_url_path) # URLデコード
        dxf_filename = os.path.join(settings.BASE_DIR, decoded_url_path.lstrip('/'))
        # print(dxf_filename)
        #      ↑ dxfファイルのフルパス
        
        # ファイルをバイトデータとして読み込む
        with open(dxf_filename, 'rb') as f:
            binary = BytesIO(f.read())
        
        infra_name = Infra.objects.filter(id=pk).first()
        file_name = infra_name.title
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        dxf_filename = file_name + "_" + timestamp + ".dxf"
        
        #レスポンスをする
        return FileResponse(binary, filename = dxf_filename)
    #finally:
        #pythoncom.CoUninitialize()


def entity_extension(mtext, neighbor):
    # MTextの挿入点
    mtext_insertion = mtext.dxf.insert
    # 特定のプロパティ(Defpoints)で描かれた文字の挿入点
    neighbor_insertion = neighbor.dxf.insert
    #テキストの行数を求める
    text = mtext.plain_text()
    text_lines = text.split("\n") if len(text) > 0 else []
    # 改行で区切ったリスト数→行数
    text_lines_count = len(text_lines)
    
    # Defpointsを範囲内とするX座標範囲
    x_start = mtext_insertion[0]  # X開始位置
    x_end  = mtext_insertion[0] + mtext.dxf.width # X終了位置= 開始位置＋幅
    y_start = mtext_insertion[1] + mtext.dxf.char_height # Y開始位置
    y_end  = mtext_insertion[1] - mtext.dxf.char_height * (text_lines_count + 1) # 文字の高さ×(行数+1)
    print("～～ DXF文字情報 ～～")
    print(f"mtextテキスト:{mtext.dxf.text}\n　Dxfテキスト:{neighbor.dxf.text}")
    print(f"mtext挿入点:{mtext_insertion}\n　Def挿入点:{neighbor_insertion}\n　行数:{text_lines_count}")
    print(f"mtext文字幅:{mtext.dxf.width}\n　mtext1行当たりの文字高:{mtext.dxf.char_height}")
    print(f"X座標の取得範囲:{x_start}～{x_end}\nY座標の取得範囲:{y_start}～{y_end}")
    print("～～ DXF文字情報 ～～")
        
    # MTextの下、もしくは右に特定のプロパティで描かれた文字が存在するかどうかを判定する(座標：右が大きく、上が大きい)
    if (neighbor_insertion[0] >= x_start and neighbor_insertion[0] <= x_end):
        #y_endの方が下部のため、y_end <= neighbor.y <= y_startとする
        if (neighbor_insertion[1] >= y_end and neighbor_insertion[1] <= y_start):
            return True
    
    return False

def find_square_around_text(dxf_filename, target_text, second_target_text):
    doc = ezdxf.readfile(dxf_filename)
    msp = doc.modelspace()
    
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    output_filepath = os.path.join(desktop_path, "CAD変更一時保存.dxf")
    doc.saveas(output_filepath)
    
    text_positions = [] # 見つかったテキストの位置を格納するためのリストを作成
    extracted_text = []
    
    # MTEXTエンティティの各要素をtextという変数に代入してループ処理
    for mtext_insert_point in msp.query('MTEXT'): # モデルスペース内の「MTEXT」エンティティをすべて照会し、ループ処理
        if mtext_insert_point.dxf.text == target_text: # エンティティのテキストが検索対象のテキストと一致した場合
            text_insertion_point = mtext_insert_point.dxf.insert # テキストの挿入点(dxf.insert)を取得します。
            text_positions.append(text_insertion_point[0]) # 挿入点のX座標をリストに保存
            break

    if not text_positions: # text_positionsリストが空の場合(見つけられなかった場合)
        for mtext_insert_point in msp.query('MTEXT'): # モデルスペース内の「MTEXT」エンティティをすべて照会し、ループ処理
            if mtext_insert_point.dxf.text == second_target_text: # エンティティのテキストが検索対象のテキストと一致した場合
                text_insertion_point = mtext_insert_point.dxf.insert # テキストの挿入点(dxf.insert)を取得します。
                text_positions.append(text_insertion_point[0]) # 挿入点のX座標をリストに保存
                break
    
    # Defpointsレイヤーで描かれた正方形枠の各要素をsquare変数に代入してループ処理
    for defpoints_square in msp.query('LWPOLYLINE[layer=="Defpoints"]'): # 
        if len(defpoints_square) == 4: # 正方形(=4辺)の場合
            square_x_values = [four_points[0] for four_points in defpoints_square] # squareというリストをループして各点(point)からx座標(インデックス0の要素)を抽出
            square_min_x = min(square_x_values) # 枠の最小X座標を取得
            square_max_x = max(square_x_values) # 枠の最大X座標を取得
            
        # 文字のX座標が枠の最小X座標と最大X座標の間にあるかチェック
        # text_positionsの各要素をtext_x_positionという変数に代入してforループを処理
        for text_x_position in text_positions:
            
            # 文字の座標がDefpoints枠のX座標内にある場合
            if square_min_x <= text_x_position <= square_max_x:
                
                # print(list(square)) 4点の座標を求める 
                left_top_point = list(defpoints_square)[0][0] # 左上の座標
                right_top_point = list(defpoints_square)[1][0] # 右上の座標
                right_bottom_point = list(defpoints_square)[2][0] # 右下の座標
                left_bottom_point = list(defpoints_square)[3][0] # 左下の座標

                defpoints_max_x = max(left_top_point,right_top_point,left_bottom_point,right_bottom_point) # X座標の最大値
                defpoints_min_x = min(left_top_point,right_top_point,left_bottom_point,right_bottom_point) # X座標の最小値
                
    # 指定したX座標範囲内にあるテキストを探す
    for circle_in_text in msp.query('MTEXT'):
        if defpoints_min_x <= circle_in_text.dxf.insert.x <= defpoints_max_x and circle_in_text.dxf.layer != 'Defpoints':
        # MTextのテキストを抽出する
            text = circle_in_text.plain_text()
            x, y, _ = circle_in_text.dxf.insert
            if not text.startswith("※"):
                cad_data = text.split("\n") if len(text) > 0 else [] # .split():\nの箇所で配列に分配
                # if len(cad_data) > 0 and not text.startswith("※") and not any(keyword in text for keyword in ["×", ".", "損傷図"]):
                if len(cad_data) > 0 and not any(keyword in text for keyword in ["×", ".", "損傷図"]) and not text.endswith("径間"):
            # 改行を含むかどうかをチェックする(and "\n" in cad):# 特定の文字列で始まるかどうかをチェックする: # 特定の文字を含むかどうかをチェックする
                    related_text = "" # 見つけたMTextと関連するDefpointsレイヤの文字列を代入する変数
            # MTextの下、もしくは右に特定のプロパティ(Defpoints)で描かれた文字を探す
                    for neighbor in msp.query('MTEXT[layer=="Defpoints"]'): # DefpointsレイヤーのMTextを抽出
                    # MTextの挿入位置と特定のプロパティで描かれた文字の位置を比較する
                        if entity_extension(circle_in_text, neighbor):
                        # 特定のプロパティ(Defpoints)で描かれた文字のテキストを抽出する
                            related_text = neighbor.plain_text()
                            print(f"DXFテキストデータ:{related_text}")
                            defx, defy, _ = neighbor.dxf.insert
                        #extracted_text.append(neighbor_text)
                            break # 文字列が見つかったらbreakによりforループを終了する

                    if  len(related_text) > 0: #related_textに文字列がある＝Defpointsレイヤから見つかった場合
                        cad_data.append(related_text[:]) # cad_dataに「部材名～使用写真」までを追加
                        cad_data.append([str(x), str(y)]) # 続いてcad_dataに「MTEXT」のX,Y座標を追加
                #最後にまとめてcad_dataをextracted_textに追加する
                    extracted_text.append(cad_data[:] + [[str(defx), str(defy)]]) # extracted_textに「MTEXTとその座標」およびdefのX,Y座標を追加
                    
# << ※特記なき損傷の抽出用 ↓ >>                            
            else:
                lines = text.split('\n')# 改行でテキストを分割してリスト化
                sub_text = [[line] for line in lines]# 各行をサブリストとして持つ多重リストを構築

                pattern = r"\s[\u2460-\u3256]"# 文字列のどこかにスペース丸数字の並びがあるかをチェックする正規表現パターン
                pattern_start = r"^[\u2460-\u3256]"  # 文字列の開始が①～㉖であることをチェックする正規表現パターン
                pattern_anywhere = r"[\u2460-\u3256]"  # 文字列のどこかに①～㉖があるかをチェックする正規表現パターン
                last_found_circle_number = None  # 最後に見つかった丸数字を保持する変数

                # リストを逆順でループし、条件に応じて処理
                for i in range(len(sub_text)-1, -1, -1):  # 後ろから前にループ
                    item = sub_text[i][0]  # textリストの各サブリストの最初の要素（[0]）をitem変数に代入（地覆 ㉓-c）
                    if item.startswith("※"):
                        sub_text.remove(sub_text[i]) # 配列から除外する
                    elif re.search(pattern, item):  # itemが正規表現patternと一致している場合（スペース丸数字の並びがある）
                        last_found = item  # last_found変数にitem要素を代入（地覆 ㉓-c）
                        # print(last_found) 丸数字が付いている要素のみ出力
                    elif 'last_found' in locals():  # last_foundが定義されている（要素が代入されている）場合のみ
                        space = last_found.replace("　", " ")
                        # 大文字スペースがあれば小文字に変換
                        second = space.find(" ", space.find(" ") + 1)#10
                        # 2つ目のスペース位置まで抽出
                        sub_text[i][0] = item + last_found[second:]
                        # item:スペース丸数字の並びがない文字列
                        # last_found:スペース丸数字の並びがある文字列
                        # last_found[second:]:スペースを含めた文字列
                    elif re.match(pattern_start, item): # 文字列が①～㉖で開始するかチェック
                        last_found_circle_number = item # 丸数字の入っている要素を保持
                        sub_text.remove(sub_text[i])
                    else:
                        if last_found_circle_number is not None and not re.search(pattern_anywhere, item):
                            # 要素に丸数字が含まれておらず、直前に丸数字が見つかっている場合
                            sub_text[i][0] += " " + last_found_circle_number  # 要素の末尾に丸数字を追加

                for sub_list in sub_text:
                    # サブリストの最初の要素を取得してスペース区切りで分割
                    split_items = sub_list[0].split()
                    
                    # 分割した要素から必要なデータを取り出して新しいサブリストに格納
                    header = split_items[0] + " " + split_items[1]  # 例：'主桁 Mg0101'
                    status = split_items[2]  # 例：'①-d'
                    # photo_number = '写真番号-00'
                    # defpoints = 'defpoints'
                    
                    # 新しい形式のサブリストを作成してprocessed_listに追加
                    # new_sub_list = [header, status, photo_number, defpoints]
                    new_sub_list = [header, status]
                    extracted_text.append(new_sub_list)

                    new_sub_list.append([str(x), str(y)])
# << ※特記なき損傷の抽出用 ↑ >>
    return extracted_text

# << dxfから要素を抽出・整列してsorted_itemsに渡す >>
def create_picturelist(request, table, dxf_filename, search_title_text, second_search_title_text):
    
    extracted_text = find_square_around_text(dxf_filename, search_title_text, second_search_title_text) # 関数の定義
    # リストを処理して、スペースを追加する関数を定義
    def add_spaces(text):
        # 正規表現でアルファベットと数字の間にスペースを挿入
        return re.sub(r'(?<! )([a-zA-Z]+)(\d{2,})', r' \1\2', text)

    # 変更されたリストを保存するための新しいリスト
    new_extracted_text = []

    # 各サブリストを処理
    for sub_extracted_text in extracted_text:
        # 先頭の文字列を修正
        if " " not in sub_extracted_text[0]:
            sub_extracted_text[0] = add_spaces(sub_extracted_text[0])
        # 新しいリストに追加
        new_extracted_text.append(sub_extracted_text)

    extracted_text = new_extracted_text

    for index, data in enumerate(extracted_text):
        # 最終項目-1まで評価
        if index < (len(extracted_text) -1):
            # 次の位置の要素を取得
            next_data = extracted_text[index + 1]
            # 特定の条件(以下例だと、１要素目が文字s1,s2,s3から始まる）に合致するかチェック
            if ("月" in next_data[0] and "日" in next_data[0]) or ("/" in next_data[0]) and (re.search(r"[A-Z]", next_data[0], re.IGNORECASE) and re.search(r"[0-9]", next_data[0])):
                # 合致する場合現在の位置に次の要素を併合 and "\n" in cad
                data.extend(next_data)
                # 次の位置の要素を削除
                extracted_text.remove(next_data)
    # extracted_text = [['主桁 Mg0101', '①-d', '写真番号-00', 'defpoints'], ['主桁 Mg0902', '⑦-c', '写真番号-00', 'defpoints']]

    # それぞれのリストから文字列のみを抽出する関数(座標以外を抽出)
        def extract_text(data):
            extracted = []  # 空のリストを用意
            removed_elements = []  # バックアップ用リスト

            pattern = r'[\u2460-\u3256]'  # ⓵～㉖

            for list_item in data:  # list_item変数に要素を代入してループ処理
                # print(list_item)
                item_extracted = [item for item in list_item if isinstance(item, str)]
                
                if item_extracted:  # item_extractedが空でないことを確認
                    # 最後の要素に特定の文字が含まれているかどうかをチェック
                    contains_symbols = bool(re.search(pattern, item_extracted[-1]))

                    # '月'と'日'が最後の要素に含まれているかどうかをチェック
                    if '月' in item_extracted[-1] and '日' in item_extracted[-1] and not contains_symbols:
                        extracted.append(item_extracted[:-2])
                        # 座標や日時を削除し、removed_elementsに保存
                        removed_elements.append([item for item in list_item if item not in item_extracted[:-2]])
                    else:
                        extracted.append(item_extracted)
                        # 座標や日時を削除し、removed_elementsに保存
                        removed_elements.append([item for item in list_item if item not in item_extracted])
                else:
                    extracted.append([])
                    removed_elements.append(list_item)

            return extracted, removed_elements  # extractedの結果を関数に返す

        # 関数を使って特定の部分を抽出
        extracted_text, removed_elements = extract_text(extracted_text)

        first_item = []
        current_detail = None  # 現在処理しているdetailを追跡

        for text, removed in zip(extracted_text, removed_elements):  # 1つずつのリスト
            result_list = []
            for item in text:# 1つずつの要素
            # 各条件を個別に確認する
                space_exists = re.search(r"\s+", item) is not None # スペースを含む
                alpha_exists = re.search(r"[a-zA-Z]+", item) is not None # アルファベットを含む
                digits_exists = re.search(r"\d{2,}", item) is not None # 2桁以上の数字を含む
            
                if space_exists and alpha_exists and digits_exists:
                # 新しいdetail項目を作成し、resultsに追加します
                    current_detail = {'detail': item, 'items': []}
                    result_list.append(current_detail)
                
                else:
                # 既存のdetailのitemsに現在の項目を追加
                    if current_detail is not None:
                        current_detail['items'].append(item)
                    
        # 元の要素を結果に追加
            for elem in removed:
                result_list.append(elem)

        #print(result_list)
            first_item.append(result_list)
        
        #print(first_item)
        extracted_text = first_item
            
        sub_first_item = [] 
        for check_sub_list in extracted_text:
            first_sub_item = []
            for first_sub_list in check_sub_list:
                # 各条件を個別に確認する
                space_exists = re.search(r"\s+", str(first_sub_list)) is not None # スペースを含む
                alpha_exists = re.search(r"[a-zA-Z]+", str(first_sub_list)) is not None # アルファベットを含む
                digits_exists = re.search(r"\d{2,}", str(first_sub_list)) is not None # 2桁以上の数字を含む
                # 正規表現を使って、コンマの直後に数字以外の文字が続く場所を見つけます。
                pattern = re.compile(r',(?![0-9])')
                # print(sub_list)
        # リスト内包表記で各要素をチェックして、条件に合致する場合は置き換えを行います。
                if space_exists and alpha_exists and digits_exists and not "月" in first_sub_list:
                    # sub_list自体を文字列に変換するのではなく、detailフィールドのみを操作する
                    detail_str = first_sub_list['detail']
                    # detail_strのカンマの直後に`</br>`タグを挿入
                    processed_str = pattern.sub(",", detail_str)
                    # processed_strをMarkup関数を使ってHTML安全なマークアップに変換
                    markup_str = Markup(processed_str)
                    # markup_strをリストに包む
                    wrapped_markup_str = [markup_str]
                    # first_sub_itemリストに追加
                    first_sub_item.append(wrapped_markup_str)
            sub_first_item.append(first_sub_item)
        # [[[Markup('横桁 Cr0503')]], [[Markup('主桁 Mg0110')], [Markup('床版 Ds0101')]], [[Markup('横桁 Cr0802')]], [[Markup('排水ます Dr0102,0201')]], [[Markup('排水ます Dr0202')]], [[Markup('PC定着部 Cn1101')]], [[Markup('排水ます Dr0102,0201,0202')]]]

            def process_item(item):
                if isinstance(item, Markup):
                    item = str(item)
                
                if ',' in item:
                    sub_items = item.split(',')
                    for i, sitem in enumerate(sub_items):
                        if i > 0 and sitem[0].isnumeric():
                            before_sub_item = sub_items[i - 1]
                            before_sub_item_splitted = before_sub_item.split()
                            before_sub_item_prefix = before_sub_item_splitted[0]
                            before_sub_item_suffix = ''
                            
                            for char in before_sub_item_splitted[1]:
                                if char.isnumeric():
                                    break
                                else:
                                    before_sub_item_suffix += char
                            
                            sub_items[i] = before_sub_item_prefix + ' ' + before_sub_item_suffix + sitem
                    item = ",".join(sub_items)
                
                return item.split(',')

            first_item = []
            for sub_one in sub_first_item:
                append2 = []
                for text_items in sub_one:
                    result_items = []
                    for item in text_items:
                        processed_items = process_item(item)
                        result_items.extend(processed_items)
                    append2.append(result_items)
                first_item.append(append2)

        # << ◆損傷種類(second)の要素◆ >> 
        # リストの各要素から記号を削除する関数
        def remove_symbols(other_items):
            symbols = ['!', '[', ']', "'"]

            processed_other_items = []
            for item in other_items:
                processed_item = ''.join(c for c in item if c not in symbols)
                processed_other_items.append(processed_item)

            return processed_other_items
        
        # それ以外の要素(損傷名)を抽出
        pattern = r'[\u2460-\u2473\u3251-\u3256].*-[a-zA-Z]' # 丸数字とワイルドカードとアルファベット
        second_items = []
        for second_sub_list in extracted_text:
            filtered_sub_list = []
            for damage_item in second_sub_list:
                if 'items' in damage_item:
                # sub_list自体を文字列に変換するのではなく、detailフィールドのみを操作する
                    detail_damage = damage_item['items']
                    for split_detail_damage in detail_damage:
                        if "," in split_detail_damage:
                            join_detail_damage = ""
                            middle_damage = split_detail_damage.split(",")
                            join_detail_damage = middle_damage
                        else:
                            join_detail_damage = detail_damage
                            
                    filtered_sub_list.append(join_detail_damage)
            second_items.append(filtered_sub_list)

        third_items = []
        bottom_item = []
        damage_coordinate = []
        picture_coordinate = []
        for other_sub_list in extracted_text:
            list_count = sum(isinstance(item, list) for item in other_sub_list) # リストの中にリストがいくつあるか数える
            
            if list_count == 2: # 座標が2つのとき=Defpointsが存在するとき
                bottom_item.append(other_sub_list[-3]) # 最後から3番目の要素を抽出（写真番号-00）
                third_items.append(other_sub_list[-4]) # 最後から4番目の要素を抽出（Defpoints）
                damage_coordinate.append(other_sub_list[-2])
                picture_coordinate.append(other_sub_list[-1])
            else: # Defpointsがない時
                bottom_item.append("") # bottom:写真番号なし
                third_items.append(None) # third:Defpointsなし
                damage_coordinate.append(other_sub_list[-1]) # damage:
                picture_coordinate.append(None) # picture:写真指定なし
        #print(other_sub_list)
        
        result_items = []# 配列を作成
        for item in bottom_item:# text_itemsの要素を1つずつitem変数に入れてforループする
            if ',' in item:# 要素の中にカンマが含まれている場合に実行
                pattern = r',(?![^(]*\))'
                sub_items = re.split(pattern, item)# カンマが含まれている場合カンマで分割
                extracted_item = []# 配列を作成
                for item in sub_items:# bottom_itemの要素を1つずつitem変数に入れてforループする
                    for p in range(len(item)):#itemの文字数をiに代入
                        if "A" <= item[p].upper() <= "Z" and p < len(item) - 1 and item[p+1].isnumeric():#i文字目がアルファベットかつ、次の文字が数字の場合
                            extracted_item.append(item[:p+1]+"*/*"+item[p+1:])# アルファベットと数字の間に*/*を入れてextracted_itemに代入
                            break
                join = ",".join(extracted_item)# 加工した内容をカンマ区切りの１つの文字列に戻す
                result_items.append(join)# result_itemsに格納

            else:# ifがfalseの場合(カンマが含まれていない場合)
                non_extracted_item = ''
                for j in range(len(item)):
                    if "A" <= item[j].upper() <= "Z" and j < len(item) - 1 and item[j+1].isnumeric():#i文字目がアルファベットかつ、次の文字が数字の場合
                        non_extracted_item = item[:j+1]+"*/*"+item[j+1:]#アルファベットまでをextracted_itemに代入
                    elif non_extracted_item == '':
                        non_extracted_item = item
                result_items.append(non_extracted_item)

        def remove_parentheses_from_list(last):
            pattern = re.compile(r"\([^()]*\)")
            result = [pattern.sub("", string) for string in last]
            return result

        last_item = remove_parentheses_from_list(result_items)

        damage_table = []  # 空のリストを作成
        # table_instance = Table.objects.filter(infra=pk).first()
        # print(f"写真パス:{table_instance.infra.article.ファイルパス}")
        # print(f"橋梁名:{table_instance.infra.title}")
        # first_itemの要素の数だけループ処理
        for i in range(len(first_item)):
            try:
                third = third_items[i]
            except IndexError:
                third = None
            
            # ['NON-a', '9月7日 S404', '9月7日 S537', '9月8日 S117,9月8日 S253']
            if len(last_item)-1 < i:
                break

            if isinstance(last_item[i], list):
                continue
            else:
                # 組み合わせを収集するリスト
                replacements = []

                # name_entriesの取得 NameEntry.objects.all()
                # tableにarticleが紐付いているため、そこから取得(tableのinfraのarticle(id))
                name_entries = NameEntry.objects.filter(article = table.infra.article)
                # print(name_entries)

                # 置換情報を収集する
                for name_entry in name_entries:
                    replacements.append((name_entry.alphabet, name_entry.name))
                replacements.append((" ", "　"))
                # print(f'replacements: {replacements}')
                # 置換リストをキーの長さで降順にソート
                sorted_replacements = sorted(replacements, key=lambda x: len(x[0]), reverse=True)

                # 置換関数を定義
                def replace_all(text, replacements):
                    return reduce(lambda acc, pair: acc.replace(pair[0], pair[1]), replacements, text)
                
                name_item = replace_all(last_item[i], sorted_replacements)
                #name_item = last_item[i].replace("S", "佐藤").replace("H", "濵田").replace(" ", "　")
            # name_item に格納されるのは 'NON-a', '9月7日 佐藤*/*404', '9月7日 佐藤*/*537', '9月8日 佐藤*/*117,9月8日 佐藤*/*253'のいずれか
            
            pattern = r',(?![^(]*\))'
            dis_items = re.split(pattern, name_item)#「9月8日 S*/*117」,「9月8日 S*/*253」
            # コンマが付いていたら分割
            
            time_result = []
            current_date = ''  # 現在の日付を保持する変数
            for time_item in dis_items:
                #print(f"このデータは：{time_item}")
                # 先頭が数字で始まるかチェック（日付として扱えるか）
                if re.match(r'^\d', time_item):
                    current_date = re.match(r'^\d+月\d+日', time_item).group(0)  # 日付を更新
                    time_result.append(time_item)  # 日付がある項目はそのまま追加
                else:
                    # 日付がない項目は、現在の日付を先頭に追加
                    time_result.append(''.join([current_date, '　', time_item]))

            sub_dis_items = ['infra/static/infra/img/' + item + ".jpg" for item in time_result]
            # sub_dis_items = [table_instance.infra.article.ファイルパス + "\\" + table_instance.infra.title + "*\\" + item + ".jpg" for item in time_result]
            # table_instance.infra.article.ファイルパス # C%3A%5CUsers%5Cdobokuka4%5CDesktop/(件名なし)/案件名/写真
            # table_instance.infra.title # サンプル橋
            # 「C:\work\django\myproject\program\Infraproject\」+「infra\static\infra\img\」+「9月7日　佐藤　地上」
            # dis_itemsの要素の数だけ、分割した各文字の先頭に「infra/static/infra/img/」各文字の後ろに「.jpg」を追加
            # ['infra/static/infra/img/9月8日 S*/*117.jpg', 'infra/static/infra/img/9月8日 S*/*253.jpg']
            # print(f"このデータは：{sub_dis_items}")
            photo_paths = []
            # photo_pathsリストを作成
            for item in sub_dis_items:
                # decoded_item = urllib.parse.unquote(item) # デコード
                # normalized_item = decoded_item.replace('/', '\\')
                # print(f"decoded_item:{decoded_item}")
                #print(f"item：{item}")
                # sub_photo_paths = glob.glob(normalized_item)
                sub_photo_paths = glob.glob(item)
                photo_paths.extend(sub_photo_paths)
                # photo_pathsリストにsub_photo_pathsを追加
            
            if len(photo_paths) > 0:# photo_pathにはリストが入るため、[i]番目の要素が0より大きい場合
                picture_urls = [''.join(photo_path).replace('infra/static/', '') for photo_path in photo_paths]
                # picture_urls = [''.join(photo_path).replace('infra/static/', '').replace('infra/img\\', '') for photo_path in photo_paths]
                # photo_pathsの要素の数だけphoto_pathという変数に代入し、forループを実行
                # photo_pathという1つの要素の'infra/static/'を空白''に置換し、中間文字なしで結合する。
                # picture_urlsという新規配列に格納する。
                # print(f"photo_paths：{picture_urls}")
            else:# それ以外の場合
                picture_urls = None
                #picture_urlsの値は[None]とする。
            
    # << ◆写真メモを作成するコード◆ >>

            bridge_damage = [] # すべての"bridge"辞書を格納するリスト

            bridge = {
                "parts_name": first_item[i],
                "damage_name": second_items[i] if i < len(second_items) else None  # second_itemsが足りない場合にNoneを使用
            }
            bridge_damage.append(bridge)

    # << ◆1つ1つの部材に対して損傷を紐付けるコード◆ >>
            first_element = bridge_damage[0]

            # 'first'キーの値にアクセス
            first_value = first_element['parts_name']

            first_and_second = []
            #<<◆ 部材名が1種類かつ部材名の要素が1種類の場合 ◆>>
            if len(first_value) == 1: # 部材名称が1つの場合
                if len(first_value[0]) == 1: # 要素が1つの場合
                    # カッコを1つ減らすためにリストをフラットにする
                    flattened_first = [first_buzai_item for first_buzai_sublist in first_value for first_buzai_item in first_buzai_sublist]
                    first_element['parts_name'] = flattened_first
                    # 同様に 'second' の値もフラットにする
                    second_value = first_element['damage_name']
                    flattened_second = [second_name_item for second_name_sublist in second_value for second_name_item in second_name_sublist]
                    first_element['damage_name'] = flattened_second

                    first_and_second.append(first_element)
                    #print(first_and_second) # [{'first': ['排水管 Dp0102'], 'second': ['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']}]

                #<<◆ 部材名が1種類かつ部材名の要素が複数の場合 ◆>>
                else: # 別の部材に同じ損傷が紐付く場合
                        # 元のリストから各要素を取得
                    for first_buzai_item in bridge_damage:
                        #print(item)
                        before_first_elements = first_buzai_item['parts_name'][0]  # ['床版 Ds0201', '床版 Ds0203']
                        first_elements = []

                        for first_buzai_second_name in before_first_elements:
                            if "～" in first_buzai_second_name:

                                first_step = first_buzai_second_name

                                if " " in first_step:
                                    # 部材記号の前にスペースが「含まれている」場合
                                    first_step_split = first_step.split()

                                else:
                                    # 部材記号の前にスペースが「含まれていない」場合
                                    first_step_split = re.split(r'(?<=[^a-zA-Z])(?=[a-zA-Z])', first_step) # アルファベット以外とアルファベットの並びで分割
                                    first_step_split = [kara for kara in first_step_split if kara] # re.split()の結果には空文字が含まれるので、それを取り除く

                                # 正規表現
                                number = first_step_split[1]
                                # マッチオブジェクトを取得
                                number_part = re.search(r'[A-Za-z]*(\d+～\d+)', number).group(1)

                                one = number_part.find("～")

                                start_number = number_part[:one]
                                end_number = number_part[one+1:]

                                # 最初の2桁と最後の2桁を取得
                                start_prefix = start_number[:2]
                                start_suffix = start_number[2:]
                                end_prefix = end_number[:2]
                                end_suffix = end_number[2:]

                                # 「主桁 Mg」を抽出
                                prefix_text = first_step_split[0] + " " + re.match(r'[A-Za-z]+', number).group(0)

                                # 決められた範囲内の番号を一つずつ追加
                                for prefix in range(int(start_prefix), int(end_prefix)+1):
                                    for suffix in range(int(start_suffix), int(end_suffix)+1):
                                        number_items = "{}{:02d}{:02d}".format(prefix_text, prefix, suffix)
                                        first_elements.append(number_items)
                            else:
                                first_elements.append(first_buzai_second_name)
                        
                        
                        second_elements = first_buzai_item['damage_name'][0]  # ['⑦剥離・鉄筋露出-d']

                        
                        # first の要素と second を一対一で紐付け
                        for first_buzai_second_name in first_elements:
                            first_and_second.append({'parts_name': [first_buzai_second_name], 'damage_name': second_elements})

                #print(first_and_second) # [{'first': '床版 Ds0201', 'second': '⑦剥離・鉄筋露出-d'}, {'first': '床版 Ds0203', 'second': '⑦剥離・鉄筋露出-d'}]

            #<<◆ 部材名が複数の場合 ◆>>
            else:
                for double_item in bridge_damage:
                    first_double_elements = double_item['parts_name'] # [['支承本体 Bh0101'], ['沓座モルタル Bm0101']]
                    second_double_elements = double_item['damage_name'] # [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e'], ['⑦剥離・鉄筋露出-c']]
                    
                    for break_first, break_second in zip(first_double_elements, second_double_elements):
                        first_and_second.append({'parts_name': break_first, 'damage_name': break_second})

            for damage_parts in bridge_damage:
                # print(damage_parts)
                if isinstance(damage_parts["damage_name"], list):  # "second"がリストの場合
                    filtered_second_items = []
                    for sublist in damage_parts["damage_name"]:
                        if isinstance(sublist, list):  # サブリストがリストである場合
                            if any(item.startswith('①') for item in sublist) and any(item.startswith('⑤') for item in sublist):
                                # ⑤で始まる要素を取り除く
                                filtered_sublist = [item for item in sublist if not item.startswith('⑤')]
                                filtered_second_items.append(filtered_sublist)
                            else:
                                filtered_second_items.append(sublist)
                        else:
                            filtered_second_items.append([sublist])
                    
                    # フィルタリング後のsecond_itemsに対して置換を行う                
                    #pavement_items = {"first": first_item[i], "second": filtered_second_items}
                        
            combined_list = []
            if damage_parts["damage_name"] is not None:
                combined_second = filtered_second_items #if i < len(updated_second_items) else None
            else:
                combined_second = None
            
            combined = {"parts_name": first_item[i], "damage_name": combined_second}
            combined_list.append(combined)
            request_list = combined_list[0]
            # <<◆ secondの多重リストを統一させる ◆>>
            try:
                # データを取得する
                check_request_list = request_list['parts_name'][1]

                # 条件分岐
                if isinstance(check_request_list, list):
                    request_list
                    #print(request_list)
                    
            except (KeyError, IndexError) as e:
                # KeyError や IndexError の例外が発生した場合の処理

                # secondの多重リストをフラットなリストに変換
                flat_list = [item for sublist in request_list['damage_name'] for item in sublist]
                # フラットなリストを再びサブリストに変換して格納
                request_list['damage_name'] = [flat_list]
                # 完成目標の確認
                
                test = request_list['damage_name'][0]

            # 先頭が文字（日本語やアルファベットなど）の場合
            def all_match_condition(lst):
                """
                リスト内のすべての項目が特定条件に一致するか確認します。
                ただし、空のリストの場合、Falseを返します。
                """
                # 空のリストの場合は False を返す
                if not lst:
                    return False
                
                pattern = re.compile(r'\A[^\W\d_]', re.UNICODE)
                return all(pattern.match(item) for item in lst)

            if all_match_condition(test):
                request_list
            else:
                request_list['damage_name'] = [request_list['damage_name']]

            #<< ◆損傷メモの作成◆ >>
            replacement_patterns = {
                "①腐食(小小)-b": "腐食", # 1
                "①腐食(小大)-c": "拡がりのある腐食",
                "①腐食(大小)-d": "板厚減少を伴う腐食",
                "①腐食(大大)-e": "板厚減少を伴う拡がりのある腐食",
                "③ゆるみ・脱落-c": "ボルト、ナットにゆるみ・脱落（●本中●本）",
                "③ゆるみ・脱落-e": "ボルト、ナットにゆるみ・脱落（●本中●本）", # 3
                "④破断-e": "鋼材の破断", # 4
                "⑤防食機能の劣化(分類1)-e": "点錆", # 5
                "⑥ひびわれ(小小)-b": "最大幅0.0mmのひびわれ", # 6
                "⑥ひびわれ(小大)-c": "最大幅0.0mmかつ間隔0.5m未満のひびわれ",
                "⑥ひびわれ(中小)-c": "最大幅0.0mmのひびわれ",
                "⑥ひびわれ(中大)-d": "最大幅0.0mmかつ間隔0.5m未満のひびわれ",
                "⑥ひびわれ(大小)-d": "最大幅0.0mmのひびわれ",
                "⑥ひびわれ(大大)-e": "最大幅0.0mmかつ間隔0.5m未満のひびわれ",
                "⑦剥離・鉄筋露出-c": "コンクリートの剥離", # 7
                "⑦剥離・鉄筋露出-d": "鉄筋露出",
                "⑦剥離・鉄筋露出-e": "断面減少を伴う鉄筋露出",
                "⑧漏水・遊離石灰-c": "漏水", # 8
                "⑧漏水・遊離石灰-d": "遊離石灰",
                "⑧漏水・遊離石灰-e": "著しい遊離石灰・泥や錆汁の混入を伴う漏水",
                "⑨抜け落ち-e": "コンクリート塊の抜け落ち", # 9
                "⑪床版ひびわれ-b": "最大幅0.0mmの1方向ひびわれ",
                "⑪床版ひびわれ-c": "最大幅0.0mmの1方向ひびわれ",
                "⑪床版ひびわれ-d": "最大幅0.0mmの1方向ひびわれ",
                "⑪床版ひびわれ-e": "最大幅0.0mmの角落ちを伴う1方向ひびわれ", # 11
                "⑫うき-e": "コンクリートのうき", # 12
                "⑮舗装の異常-c": "最大幅0.0mmのひびわれ",
                "⑮舗装の異常-e": "最大幅0.0mmのひびわれ・舗装の土砂化", # 15
                "⑯定着部の異常-c": "定着部の損傷。",
                "⑯定着部の異常(分類2)-e": "定着部の著しい損傷", # 16
                "⑳漏水・滞水-e": "漏水・滞水", # 20
                "㉓変形・欠損-c": "変形・欠損", # 23
                "㉓変形・欠損-e": "著しい変形・欠損",
                "㉔土砂詰まり-e": "土砂詰まり", # 24
            }

            def describe_damage(unified_request_list):
                described_list = []
                
                for damage in unified_request_list:
                    if damage in replacement_patterns:
                        described_list.append(replacement_patterns[damage])
                    elif damage.startswith('⑰'):
                        match = re.search(r'(?<=:)(.*?)(?=\)-e)', damage)
                        if match:
                            described_list.append(match.group(1))
                    else:
                        pattern = r'[\u3248-\u3257](.*?)-'
                        match = re.search(pattern, damage)
                        if match:
                            described_list.append(match.group(1))
                        else:
                            described_list.append(damage)  # フォールバックとしてそのまま返す
                return ','.join(described_list)

            # 各ケースに対して出力を確認:
            def generate_report(unified_request_list):
                primary_damages = []
                processed_related_damages = []
                #print(f"unified_request_list：{unified_request_list}")
                first_items = unified_request_list['parts_name']
                #print(first_items) # [['支承本体 Bh0101'], ['沓座モルタル Bm0101']]
                second_items = unified_request_list['damage_name']
                #print(second_items) # [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e'], ['⑦剥離・鉄筋露出-c']]
                primary_damages_dict = {}

                for first_item, second_item in zip(first_items, second_items):
                    element_names = [f.split()[0] for f in first_item] # カッコ内の要素について、スペースより前を抽出
                    #print(f"element_names：{element_names}") # ['支承本体'], ['沓座モルタル']
                    damage_descriptions = describe_damage(second_item) # 辞書で置換
                    #print(f"damage_descriptions：{damage_descriptions}") # 腐食,点錆, 剥離
                    
                    if len(element_names) == 1: # ['主桁', '横桁', '対傾構']：これはだめ
                        primary_damages.append(f"{element_names[0]}に{damage_descriptions}が見られる。")
                        #print(f"primary_damages：{primary_damages}") # ['支承本体に腐食,点錆が見られる。', '沓座モルタルに剥離が見られる。']
                    else:
                        element_names = list(dict.fromkeys(element_names))            
                        joined_elements = "および".join(element_names[:-1]) + "," + element_names[-1]
                        if joined_elements.startswith(","):
                            new_joined_elements = joined_elements[1:]
                        else:
                            new_joined_elements = joined_elements
                        primary_damages.append(f"{new_joined_elements}に{damage_descriptions}が見られる。")

                    for elem in first_item:
                        primary_damages_dict[elem] = second_item[:]

                primary_description = "また".join(primary_damages)
                    
                for elem_name, elem_number in zip(first_items, second_items): # 主桁 Mg0101
                    # リストをフラットにする関数
                    def flatten_list(nested_list):
                        return [item for sublist in nested_list for item in sublist]
                    
                    # 辞書から 'first' と 'second' の値を取り出す
                    first_list = request_list['parts_name']
                    second_list = request_list['damage_name']

                    # 'first' の要素数を数える
                    first_count = sum(len(sublist) for sublist in first_list)

                    # 'second' の要素数を数える
                    second_count = sum(len(sublist) for sublist in second_list)
                    # フラットにしたリストを比較
                    if flatten_list(first_items) != elem_name and flatten_list(second_items) != elem_number:
                        sub_related_damages = []
                        for first_item in first_items:
                            for elem in first_item:
                                if elem in primary_damages_dict:
                                    formatted_damages = ",".join(list(dict.fromkeys(primary_damages_dict[elem])))
                                    sub_related_damages.append(f"{elem}:{formatted_damages}")
                                    #print(f"sub_related_damages：{sub_related_damages}") # ['支承本体 Bh0101:①腐食(小小)-b,⑤防食機能の劣化(分類1)-e', '沓座モルタル Bm0101:⑦剥離・鉄筋露出-c']

                        # 処理後のリストを格納するための新しいリスト
                        second_related_damages = []

                        # リスト内の各要素をループする
                        for i, damage in enumerate(sub_related_damages):
                            # コロンの位置を取得
                            colon_index = damage.find(":")
                            
                            if colon_index != -1:
                                if i == 0:
                                    # 1番目の要素の場合
                                    parts = damage.split(',')
                                    
                                    if len(parts) > 1:
                                        first_damage = parts[0].split(':')[0]
                                        after_damage = ':' + parts[1].strip()
                                        create_damage = first_damage + after_damage
                                        second_related_damages.append(create_damage)

                                else:
                                    # 2つ目以降の要素の場合
                                    parts = damage.split(',')
                                    second_related_damages.append(damage)
                                    

                        # 処理後のリストを格納するための新しいリスト
                        processed_related_damages = []
                        #print(f"second_related_damages：{second_related_damages}")
                        for damage in second_related_damages:
                            colon_index = damage.find(":")
                            if colon_index != -1:
                                before_colon_part = damage[:colon_index].strip()
                                after_colon_part = damage[colon_index + 1:].strip()
                                #print(f"damage[colon_index + 1:]：{damage}")
                                if before_colon_part and after_colon_part:
                                    processed_damage = f"{before_colon_part}:{after_colon_part}"
                                    processed_related_damages.append(processed_damage)
                        #print(f"after_colon_part：{processed_related_damages}")
                        
                    elif first_count < 2 and second_count < 2: # {'first': [['横桁 Cr0803']], 'second': [['⑦剥離・鉄筋露出-d']]}
                        None
                    elif first_count > 1 and second_count < 2: # {'first': [['床版 Ds0201', '床版 Ds0203']], 'second': [['⑦剥離・鉄筋露出-d']]}
                        first_items_from_first = first_item[1:]
                        related_damage_list = ','.join(first_items_from_first)# カンマ区切りの文字列に結合
                        related_second_item = ','.join(second_item)
                        processed_related_damages.append(f"{related_damage_list}:{related_second_item}")
                    elif first_count < 2 and second_count > 1: # {'first': [['横桁 Cr0503']], 'second': [['⑦剥離・鉄筋露出-d', '⑰その他(分類6:施工不良)-e']]}
                        second_items_from_second = second_item[1:]
                        related_damage_list = ','.join(second_items_from_second)# カンマ区切りの文字列に結合
                        processed_related_damages.append(f"{','.join(elem_name)}:{related_damage_list}")
                    else:#  len(elem_name) > 1 and len(elem_number) > 1: # {'first': [['排水管 Dp0101', '排水管 Dp0102']], 'second': [['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']]}
                        related_damage_list = ','.join(second_item)
                        processed_related_damages.append(f"{','.join(elem_name)}:{related_damage_list}")


                related_description = ""
                if processed_related_damages:
                    related_description = "\n【関連損傷】\n" + ", ".join(processed_related_damages)

                return f"{primary_description} {related_description}".strip()

            combined_data = generate_report(request_list)
            #print(combined_data)
            # print(f"picture_urls：{picture_urls}")
            
                    # \n文字列のときの改行文字
            items = {'parts_name': first_item[i], 'damage_name': second_items[i], 'join': first_and_second, 
                     'picture_number': third, 'this_time_picture': picture_urls, 'last_time_picture': None, 'textarea_content': combined_data, 
                     'damage_coordinate': damage_coordinate[i], 'picture_coordinate': picture_coordinate[i]}
            print(f"items：{items}")
            damage_table.append(items)

        #優先順位の指定
        order_dict = {"主桁": 1, "横桁": 2, "床版": 3, "PC定着部": 4, "橋台[胸壁]": 5, "橋台[竪壁]": 6, "支承本体": 7, "沓座モルタル": 8, "防護柵": 9, "地覆": 10, "伸縮装置": 11, "舗装": 12, "排水ます": 13, "排水管": 14}
        order_number = {"None": 0, "①": 1, "②": 2, "③": 3, "④": 4, "⑤": 5, "⑥": 6, "⑦": 7, "⑧": 8, "⑨": 9, "⑩": 10, "⑪": 11, "⑫": 12, "⑬": 13, "⑭": 14, "⑮": 15, "⑯": 16, "⑰": 17, "⑱": 18, "⑲": 19, "⑳": 20, "㉑": 21, "㉒": 22, "㉓": 23, "㉔": 24, "㉕": 25, "㉖": 26}
        order_lank = {"a": 1, "b": 2, "c": 3, "d": 4, "e": 5}
                
        # <<◆ リストの並び替え ◆>>
        def sort_key_function(sort_item):
            first_value = sort_item['parts_name'][0][0] # firstキーの最初の要素
            #print(first_value) # 主桁 Mg0901

            if " " in first_value:
                # 部材記号の前にスペースが「含まれている」場合
                first_value_split = first_value.split()
                #print(first_value_split) # ['主桁', 'Mg0901']
            else:
                # 部材記号の前にスペースが「含まれていない」場合
                first_value_split = re.split(r'(?<=[^a-zA-Z])(?=[a-zA-Z])', first_value) # アルファベット以外とアルファベットの並びで分割
                first_value_split = [x for x in first_value_split if x] # re.split()の結果には空文字が含まれるので、それを取り除く
                #print(f"first_value_split：{first_value_split}") # ['主桁', 'Mg0901']

            first_name_key = order_dict.get(first_value_split[0], float('inf'))
            #print(first_name_key) # 1
            if "～" in first_value_split[1]:
                match = re.search(r'[A-Za-z]+(\d{2,})(\D)', first_value_split[1])
                if match:
                    first_number_key = int(match.group(1))
            else:
                first_number_key = int(first_value_split[1][2:])
            #print(first_number_key) # 901

            if sort_item['damage_name'][0][0]:  # `second`キーが存在する場合
                second_value = sort_item['damage_name'][0][0] # secondキーの最初の要素
                #print(second_value) # ⑰その他(分類6:異物混入)-e
                second_number_key = order_number.get(second_value[0], float('inf'))  # 先頭の文字を取得してorder_numberに照らし合わせる
                #print(second_number_key) # 17
                second_lank_key = order_lank.get(second_value[-1], float('inf'))  # 末尾の文字を取得してorder_lankに照らし合わせる
                #print(second_lank_key) # 5
            else:
                second_number_key = float('inf')
                second_lank_key = float('inf')
                    
            return (first_name_key, first_number_key, second_number_key, second_lank_key)

        sorted_items = sorted(damage_table, key=sort_key_function)
        # print(f"sorted_items：{sorted_items}")
    return sorted_items

# << 旗揚げの修正 >>
def edit_report_data(request, damage_pk, pk, table_pk):
    print(f"damage_pk={damage_pk} pk={pk} table_pk={table_pk}")
    report_data = get_object_or_404(FullReportData, pk=damage_pk)
    if request.method == "POST":
        points = request.POST.get("coords").split(",")
        coords = [float(points[0]), float(points[1])]
        print(f"変更前coords:{coords}")
        
        # DXFファイルの更新処理
        def find_square_around_text(dxf_filename, target_text, second_target_text):
            doc = ezdxf.readfile(dxf_filename)
            msp = doc.modelspace()
            print(f"DOC:{doc}")
            print(f"MSP:{msp}")
            print(f"dxf_filename:{dxf_filename}")
            # 座標の一致を確認するための許容誤差
            epsilon = 0.001

            for entity in msp:
                if entity.dxftype() in {'TEXT', 'MTEXT'}:
                    x, y, _ = entity.dxf.insert

                    if abs(float(x) - float((points[0]))) < epsilon and abs(float(y) - float(points[1])) < epsilon:
                        print(f"変更前ENTITY:{entity}")
                        print(f"変更前DXFテキスト:{entity.dxf.text}")
                        return entity.dxf.text
        
        infra_name = Infra.objects.filter(pk=pk).first()
        article_name = infra_name.article
        print(f"Infra:{infra_name}/Article:{article_name}")
        table = Table.objects.get(infra=infra_name, article=article_name)
        print(f"変更TABLE:{table}") # Table object (5)
        table_number = table.id
        print(f"TABLEのID:{table_number}")
        table = Table.objects.filter(pk=table_pk).first()
        print("～～～")
        print(f"TABLE:{pk}")
        print(f"変更TABLE:{table}") # Table object (5)
        print(f"変更後coords:{coords}") # [525003.839727268, 214191.031706055]
        if not table:
            print(f"変更TABLE:データなし")
        
        encoded_url_path = table.dxf.url
        decoded_url_path = urllib.parse.unquote(encoded_url_path) # URLデコード
        dxf_filename = os.path.join(settings.BASE_DIR, decoded_url_path.lstrip('/'))
        
        current_text = find_square_around_text(dxf_filename, coords[0], coords[1])

        return JsonResponse({"status": "success", 'current_text': current_text})

    return render(request, 'infra/bridge_table.html', {'report_data': report_data})

@csrf_exempt
def edit_send_data(request, damage_pk, pk, table_pk):
    print(f"修正対象：damage_pk={damage_pk} pk={pk} table_pk={table_pk}")
    report_data = get_object_or_404(FullReportData, pk=damage_pk)
    if request.method == "POST":
        data = json.loads(request.body)
        points = data.get('coords')
        new_text = data.get('new_text')
        print(f"変更points:{points}")
        print(f"変更new_text:{new_text}")
        def find_square_around_text(dxf_filename, new_text):
            doc = ezdxf.readfile(dxf_filename)
            msp = doc.modelspace()
            epsilon = 0.001
                
            for entity in msp:
                if entity.dxftype() in {'TEXT', 'MTEXT'}:
                    insert_point = entity.dxf.insert
                    x,y = insert_point.x, insert_point.y
                    x_points, y_points = points.split(',')
                    print(f"変更    座標:{entity.dxf.text}--{float(x)}/{float(x_points)}/{float(y)}/{float(y_points)}//{points}")
                    if abs(float(x) - float(x_points)) < epsilon and abs(float(y) - float(y_points)) < epsilon: # 座標が一致した場合(誤差:epsilon)
                        entity.dxf.text = new_text  # 新しいテキストに置き換え
                        print(f"変更前文字:{new_text}")
                        print(f"変更後文字:{entity.dxf.text}")
                        
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
            output_filepath = os.path.join(desktop_path, "CAD変更.dxf")
            doc.saveas(output_filepath)
            print(f"変更完了:{new_text}")
            return new_text
                              
        table = Table.objects.filter(pk=table_pk).first()
        
        encoded_url_path = table.dxf.url
        decoded_url_path = urllib.parse.unquote(encoded_url_path)  # URLデコード
        dxf_filename = os.path.join(settings.BASE_DIR, decoded_url_path.lstrip('/'))

        current_text = find_square_around_text(dxf_filename, new_text)

        return JsonResponse({"status": "success", 'current_text': current_text})

    return render(request, 'infra/bridge_table.html', {'report_data': report_data})