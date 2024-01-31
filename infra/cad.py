# @title デフォルトのタイトル テキスト
# C:\work\django\myproject\myvenv\Scripts\python.exe
import ezdxf
import openpyxl
from ezdxf.entities.mtext import MText
from ezdxf.entities.text import Text
import pandas as pd

dxf = ezdxf.readfile(R'C:\work\django\myproject\myvenv\Infraproject\uploads\5_大久保歩道橋.dxf') # ファイルにアップロードしたdxfファイル名

cad_read = []
for entity in dxf.entities:
    if type(entity) is MText: # or type(entity) is Text: MTextとTextが文字列を表す(https://ymt-lab.com/post/2021/ezdxf-read-dxf-file/)
        cad =  entity.plain_text() # plain_text(読める文字)に変換
        cad_data = cad.split("\n") if len(cad) > 0 else [] # .split():\nの箇所で改行
        if len(cad_data) > 0:
            cad_read.append(cad_data)
print( cad_read )

df = pd.DataFrame( cad_read )
html = df.to_html(index=False)

with open("pandas_table01.html", "a", encoding="utf-8") as file:# "a":末尾に追記
    file.write(html)

# workbook = openpyxl.Workbook() # 新規ファイルの作成
# sheet = workbook.active # アクティブなシートの選択

# row = 1  # 出力する行数

# for entity in dxf.entities: # 上と同じ
#     if type(entity) is MText: # 上と同じ
#         text = entity.plain_text() # textに格納
#         sheet.cell(row=row, column=1).value = text #sheet.cell():セルの選択、row:行、column:列 → A列を1行ずつ選択
#         row += 1 # 一段下げる

# workbook.save('歩道橋.xlsx')# 保存するエクセル名