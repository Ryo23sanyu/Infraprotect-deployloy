import os
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO

def extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))

    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 「写真番号」と書かれているセルを探す
    photo_number_positions = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "写真番号":
                photo_number_positions.append((cell.row, cell.column))

    # 写真番号セルから右方向に移動した位置の値を取得するための辞書作成
    photo_name_dict = {}
    for p_row, p_col in photo_number_positions:
        for offset in range(1, 10):  # 最大10列までチェックする
            photo_name_cell = sheet.cell(row=p_row, column=p_col + offset)
            if photo_name_cell.value:
                photo_name_dict[(p_row, p_col)] = photo_name_cell.value
                break  # 最初に見つかったもののみを使用する

    # 画像を保存する
    for img_idx, (row, col, img) in enumerate(images):
        try:
            # 画像データを取り出し
            img_stream = BytesIO(img._data())
            img_data = img_stream.read()

            # 前回の写真番号を取得
            nearest_photo_number = None
            for key in photo_name_dict:
                if key[0] <= row and (nearest_photo_number is None or key[0] > nearest_photo_number[0]):
                    nearest_photo_number = key

            if nearest_photo_number:
                file_name = photo_name_dict[nearest_photo_number]
            else:
                file_name = f'image_{img_idx + 1}'

            # 画像を保存するパス
            image_file_path = os.path.join(output_folder, f"画像：{file_name}.jpg")

            # 画像データを保存
            with open(image_file_path, "wb") as img_file:
                img_file.write(img_data)
                print(img_file)
                
        except Exception as e:
            print(f"Error processing image at row {row} and col {col}: {e}")

# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定

extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder)