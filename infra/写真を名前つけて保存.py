import os
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO

def extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder):
    # 保存先フォルダが存在しない場合は作成
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Excelファイルを読み込む
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb[sheet_name]

    # シート内の画像を収集
    images = []
    for img in sheet._images:
        images.append((img.anchor._from.row, img.anchor._from.col, img))

    # 行方向に並べ替え
    images.sort(key=lambda x: (x[0], x[1]))

    # 「写真番号」と書かれているセルを探す
    photo_number_positions = []
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value == "写真番号":
                photo_number_positions.append((cell.row, cell.column))

    # 写真番号セルから右方向に移動した位置の値を取得するための辞書作成
    photo_name_dict = {}
    for p_row, p_col in photo_number_positions:
        for offset in range(1, 10):  # 最大10列までチェックする
            photo_name_cell = sheet.cell(row=p_row, column=p_col + offset)
            if photo_name_cell.value:
                photo_name_dict[(p_row, p_col)] = photo_name_cell.value
                #print(photo_name_dict)
                # {(9, 4): 1, (9, 24): 2, (9, 44): '前回写真-3', (23, 4): 3, (23, 24): 4, (23, 44): 5, (46, 4): 6, (46, 24): 7, (46, 44): 8, (60, 4): 9, (60, 24): 10, (60, 44): 11, (83, 4): 12, (83, 24): 13, (83, 44): 14, (97, 4): 15, (97, 24): 16, (97, 44): 17, (120, 4): 18, (120, 24): 19, (120, 44): 20, (134, 4): 21, (134, 24): 22, (134, 44): 23}
                break  # 最初に見つかったもののみを使用する

    # 画像を保存する
    for img_idx, (row, col, img) in enumerate(images): # imagesの各アイテム (row, col, img) に対して、インデックス (img_idx) をつけてループ処理
        # row：行, col：列, img：画像データ
        try:
            # 画像データを取り出し
            img_stream = BytesIO(img._data()) # img._data()を呼び出して画像データを取得
            img_data = img_stream.read() # バイトデータの読み取り

            # 前回の写真番号を取得
            nearest_photo_number = None # 最も近い写真番号を保持するための変数 nearest_photo_numberを Noneで初期化
            for key in photo_name_dict:
                print(f"key[0] : {key[0]} <= row : {row}")
                # photo_name_dictの各キーについてループを行い、キーの最初の要素 (key[0]) が現在の行（row）より小さいか等しいかどうかを調べる

                if (
                    (key[0] <= row and key[1] <= col) and # キーの行(row)番号が現在の行番号以下、かつ列(col)番号が現在の列番号以下の場合。かつ
                    ((nearest_photo_number is None) or (key[0] > nearest_photo_number[0]) or
                     (key[0] == nearest_photo_number[0] and key[1] > nearest_photo_number[1]))
                ):
                    # key[0]: 文字列の行数、row: 画像の行数
                    # ↓　どちらかに一致すればTrue
                    # key[0] <= row（キーの行番号が現在の行番号以下）
                    # nearest_photo_numberが Noneまたは key[0]が現在の最も近い行番号より大きい
                    nearest_photo_number = key
                    # そのkeyを nearest_photo_numberに設定

            if nearest_photo_number: # nearest_photo_numberが存在する場合
                file_name = photo_name_dict[nearest_photo_number] # photo_name_dictの対応する値（ファイル名）を取得し、file_nameとして設定
                print(nearest_photo_number)

            else:
                file_name = f'image_{img_idx + 1}'

            # 画像を保存するパス
            image_file_path = os.path.join(output_folder, f"画像：{file_name}.jpg")

            # 画像データを保存
            with open(image_file_path, "wb") as img_file:
                img_file.write(img_data)
                
        except Exception as e:
            print(f"Error processing image at row {row} and col {col}: {e}")

# 使い方
excel_file = R'C:\work\django\myproject\program\Infraproject\uploads\たぬき橋.xlsx'
sheet_name = 'その１０'  # 抽出したいシート名を指定
output_folder = '前回写真'  # 保存先フォルダを指定

extract_images_and_save_with_photo_number(excel_file, sheet_name, output_folder)