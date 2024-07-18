from copy import copy
from itertools import zip_longest
import re
import openpyxl
import os
import datetime
import tempfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

def duplicate_sheet(wb, sheet_name, suffix):
    """元のシートを複製して、指定したタイトルのシートを作成"""
    original_sheet = wb[sheet_name]
    # シートを複製
    copied_sheet = wb.copy_worksheet(original_sheet)
    # コピーシートのタイトルを更新
    copied_sheet.title = f"{sheet_name}-{suffix}"
    return copied_sheet

def set_print_area_and_page_view(sheet, start_col, end_col):
    """印刷範囲と改ページプレビューを設定する"""
    # 最大行を取得
    max_row = sheet.max_row
    # 印刷範囲の設定を修正
    print_area = f"{start_col}1:{end_col}{max_row}"
    sheet.print_area = print_area

    # 改ページプレビューを有効にする
    sheet.sheet_view.view = "pageBreakPreview"
    
def hide_sheet_copy_and_paste(wb, sheet_name):
    """シートを再表示してコピーその後非表示に設定"""

    hide_sheet = wb['ページ１０']
    hide_sheet.sheet_state = 'visible'

    # コピーする行の範囲を指定します
    copy_start_row = 2
    copy_end_row = 29

    # コピーする行のデータとスタイルを保持するリストを作成します
    rows_to_copy = []
    merges_to_keep = []

    for row_idx in range(copy_start_row, copy_end_row + 1):
        row_data = []
        for cell in hide_sheet[row_idx]:
            cell_data = {
                'value': cell.value,
                'font': copy(cell.font),
                'border': copy(cell.border),
                'fill': copy(cell.fill),
                'number_format': cell.number_format,
                'protection': copy(cell.protection),
                'alignment': copy(cell.alignment)
            }
            row_data.append(cell_data)
        row_data.append(hide_sheet.row_dimensions[row_idx].height)
        rows_to_copy.append(row_data)

    # 元のシートのセル結合情報を取得
    for merge in hide_sheet.merged_cells.ranges:
        if (copy_start_row <= merge.min_row <= copy_end_row) or \
            (copy_start_row <= merge.max_row <= copy_end_row):
            merges_to_keep.append(copy(merge))
    
    action_sheet_name = 'その１０-1'
    sheet = wb[action_sheet_name]
    
    # コピー先の行を挿入します
    # A列の一番下の行番号を取得
    max_row = sheet.max_row
    while sheet['A' + str(max_row)].value is None and max_row > 0:
        max_row -= 1
    insert_at_row = max_row
    print(f"max_row：{max_row}")
    
    # シフトする行の高さを保持するリストを作成します
    heights = []
    for row_idx in range(insert_at_row, sheet.max_row + 1):
        heights.append(sheet.row_dimensions[row_idx].height)
    
    # 指定行から下の行をシフト
    sheet.insert_rows(insert_at_row, amount=(copy_end_row - copy_start_row + 1))

    # 行の高さを元に戻す
    for i, height in enumerate(heights):
        sheet.row_dimensions[insert_at_row + i + (copy_end_row - copy_start_row + 1)].height = height
    
    # コピーされた行を挿入
    for i, row_data in enumerate(rows_to_copy):
        new_row = insert_at_row + i
        for j, cell_data in enumerate(row_data[:-1]):
            cell = sheet.cell(row=new_row, column=j + 1)
            cell.value = cell_data['value']
            cell.font = cell_data['font']
            cell.border = cell_data['border']
            cell.fill = cell_data['fill']
            cell.number_format = cell_data['number_format']
            cell.protection = cell_data['protection']
            cell.alignment = cell_data['alignment']
        sheet.row_dimensions[new_row].height = row_data[-1]

    # セル結合をコピー
    for merged_range in merges_to_keep:
        new_min_row = merged_range.min_row - copy_start_row + insert_at_row
        new_max_row = merged_range.max_row - copy_start_row + insert_at_row
        new_merge_range = "{}{}:{}{}".format(
            openpyxl.utils.get_column_letter(merged_range.min_col),
            new_min_row,
            openpyxl.utils.get_column_letter(merged_range.max_col),
            new_max_row
        )
        sheet.merge_cells(new_merge_range)
        
    # 最大行を取得
    max_row = sheet.max_row
    # 印刷範囲の設定を修正
    # start_colとend_colの取得方法が不明だったため、適宜修正してください。
    start_col = "A"
    end_col = 'CD'
    print_area = f"{start_col}1:{end_col}{max_row}"
    sheet.print_area = print_area    
    
    hide_sheet.sheet_state = 'hidden'



# 元のファイルのパス（例: `10_only.xlsm`）
original_file_path = '10_only.xlsm'

# エクセルファイルを読み込む
wb = openpyxl.load_workbook(original_file_path, keep_vba=True)

# 複製元のシートを選択 (例: 'その１０')
original_sheet_name = 'その１０'
ws = wb[original_sheet_name]

# 例えば10個の複製シートを作成
num_copies = 2
start_col = 'A'  # 印刷範囲の開始列
end_col = 'CD'  # 印刷範囲の終了列
for i in range(1, num_copies + 1):
    copied_sheet = duplicate_sheet(wb, original_sheet_name, i)
    set_print_area_and_page_view(copied_sheet, start_col, end_col)
    copied_sheet['A65'] = "python" # 半角スペースの代わり（エクセルが終わった際に置換で削除）


action_sheet_name = 'その１０-1'
ws = wb[action_sheet_name] # シート名を指定

sorted_items = [
    {'parts_name': [['排水管 Dp0101']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
     ,
    {'parts_name': [['排水管 Dp0102']], 'damage_name': [['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0102'], 'damage_name': ['①腐食(小大)-c', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-32', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070486.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0102:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['547059.1990495767', '229268.8593029478'], 'picture_coordinate': ['549204.9604817769', '229256.3408485695']}
    ,
    {'parts_name': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'damage_name': [['⑳漏水・滞水-e']], 
     'join': [{'parts_name': ['橋台[胸壁] Ap0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['橋台[竪壁] Ac0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['伸縮装置 Ej0101'], 'damage_name': ['⑳漏水・滞水-e']}], 
     'picture_number': '写真番号-18', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'last_time_picture': 'media\P6100002.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'parts_name': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'damage_name': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'parts_name': ['支承本体 Bh0101'], 'damage_name': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'parts_name': ['沓座モルタル Bm0101'], 'damage_name': ['⑦剥離・鉄筋露出-c']}], 
     'picture_number': '写真番号-27', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'parts_name': [['床版 Ds0201', '床版 Ds0203']], 'damage_name': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'parts_name': ['床版 Ds0201'], 'damage_name': ['⑦剥離・鉄筋露出-d']}, {'parts_name': ['床版 Ds0203'], 'damage_name': ['⑦剥離・鉄筋露出-d']}], 
     'picture_number': None, 'this_time_picture': None, 
     'last_time_picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'parts_name': [['排水管 Dp0103']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
    ,
    {'parts_name': [['橋台[胸壁] Ap0101', '橋台[竪壁] Ac0101', '伸縮装置 Ej0101']], 'damage_name': [['⑳漏水・滞水-e']], 
     'join': [{'parts_name': ['橋台[胸壁] Ap0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['橋台[竪壁] Ac0101'], 'damage_name': ['⑳漏水・滞水-e']}, {'parts_name': ['伸縮装置 Ej0101'], 'damage_name': ['⑳漏水・滞水-e']}], 
     'picture_number': '写真番号-18', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070438.JPG'], 
     'last_time_picture': 'media\P6100020.JPG', 
     'textarea_content': '橋台[胸壁]および橋台[竪壁],伸縮装置に漏水・滞水が見られる。 \n【関連損傷】\n橋台[竪壁] Ac0101,伸縮装置 Ej0101:⑳漏水・滞水-e', 
     'damage_coordinate': ['535305.6406762057', '190342.4721676922'], 'picture_coordinate': ['537494.8440878117', '190371.7813098583']}
    ,
    {'parts_name': [['支承本体 Bh0101'], ['沓座モルタル Bm0101']], 'damage_name': [['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c'], ['⑦剥離・鉄筋露出-c']], 
     'join': [{'parts_name': ['支承本体 Bh0101'], 'damage_name': ['①腐食(小小)-b', '⑤防食機能の劣化(分類1)-e', '㉓変形・欠損-c']}, {'parts_name': ['沓座モルタル Bm0101'], 'damage_name': ['⑦剥離・鉄筋露出-c']}], 
     'picture_number': '写真番号-27', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070504.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '支承本体に腐食,変形・欠損が見られる。また沓座モルタルにコンクリートの剥離が見られる。 \n【関連損傷】\n支承本体 Bh0101:㉓変形・欠損-c, 沓座モルタル Bm0101:⑦剥離・鉄筋露出-c', 
     'damage_coordinate': ['547050.6408404222', '164535.3885015437'], 'picture_coordinate': ['549493.5416080137', '164259.8990548863']}
    ,
    {'parts_name': [['床版 Ds0201', '床版 Ds0203']], 'damage_name': [['⑦剥離・鉄筋露出-d']], 
     'join': [{'parts_name': ['床版 Ds0201'], 'damage_name': ['⑦剥離・鉄筋露出-d']}, {'parts_name': ['床版 Ds0203'], 'damage_name': ['⑦剥離・鉄筋露出-d']}], 
     'picture_number': None, 'this_time_picture': None, 
     'last_time_picture': None, 
     'textarea_content': '床版に鉄筋露出が見られる。 \n【関連損傷】\n床版 Ds0203:⑦剥離・鉄筋露出-d', 
     'damage_coordinate': ['525003.839727268', '213095.7270112425'], 'picture_coordinate': None}
    ,
    {'parts_name': [['排水管 Dp0103']], 'damage_name': [['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']], 
     'join': [{'parts_name': ['排水管 Dp0101'], 'damage_name': ['①腐食(大大)-e', '⑤防食機能の劣化(分類1)-e']}], 
     'picture_number': '写真番号-31', 'this_time_picture': ['infra\img\9月7日　佐藤　地上\P9070422.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '排水管に板厚減少を伴う拡がりのある腐食,点錆が見られる。\n【関連損傷】\n排水管 Dp0101:⑤防食機能の劣化(分類1)-e', 
     'damage_coordinate': ['538482.3557216563', '229268.8593029478'], 'picture_coordinate': ['538810.3087944178', '228910.3502713814']}
    ,
    {'parts_name': [['横桁 Cr0102']], 'damage_name': [['⑰その他(分類6:施工不良)-e']], 
     'join': [{'parts_name': ['横桁 Cr0102'], 'damage_name': ['⑰その他(分類6:施工不良)-e']}], 
     'picture_number': '写真番号-4,5', 'this_time_picture': ['infra/img\\9月7日\u3000佐藤\u3000地上\\P9070424.JPG', 'infra/img\\9月7日\u3000佐藤\u3000地上\\P9070430.JPG'], 
     'last_time_picture': None, 
     'textarea_content': '横桁に施工不良が見られる。', 
     'damage_coordinate': ['532578.7587482664', '229268.8593029478'], 'picture_coordinate': ['532985.6409545547', '228954.2335446222']}
]

# lastがNoneでないデータを残す
filtered_data = [item for item in sorted_items if item['this_time_picture'] is not None]
print(filtered_data)

# this_time_pictureの要素が複数ある場合、分割する
output_data = []
for entry in filtered_data:
    for picture in entry['this_time_picture']:
        new_entry = entry.copy()
        new_entry['this_time_picture'] = [picture]
        output_data.append(new_entry)
        
# 行の開始地点と増加するステップを定義
partsname_and_number_row = 10 # 部材名・要素番号
damagename_and_lank_row = 11 # 損傷の種類・損傷程度
picture_start_row = 13 # 損傷写真
lasttime_lank_row = 15 # 前回損傷程度
damage_memo_row = 17 # 損傷メモ
step = 14
num_positions = len(output_data) * 3  # データ数に3列分を掛けています

# 関連する列を定義
picture_columns = ["E", "AE", "BE"]
left_columns = ["I", "AI", "BI"]
right_columns = ["R", "AR", "BR"]
bottom_columns = ["T", "AT", "BT"]

# セル位置のリストを生成 ↓
partsname_cell_positions = [] # 部材名
number_cell_positions = [] # 要素番号
damagename_cell_positions = [] # 損傷の種類
lank_cell_positions = [] # 損損傷程度
picture_cell_positions = [] # 損傷写真
lasttime_lank_cell_positions = [] # 前回損傷程度
damage_memo_cell_positions = [] # 損傷メモ

for i in range(num_positions // len(picture_columns)):
    partsname_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in left_columns])
    number_cell_positions.append([f"{col}{partsname_and_number_row + i * step}" for col in right_columns])
    damagename_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in left_columns])
    lank_cell_positions.append([f"{col}{damagename_and_lank_row + i * step}" for col in right_columns])
    picture_cell_positions.append([f"{col}{picture_start_row + i * step}" for col in picture_columns])
    lasttime_lank_cell_positions.append([f"{col}{lasttime_lank_row + i * step}" for col in bottom_columns])
    damage_memo_cell_positions.append([f"{col}{damage_memo_row + i * step}" for col in bottom_columns])
    
join_partsname_cell_positions = [item for sublist in partsname_cell_positions for item in sublist]
join_number_cell_positions = [item for sublist in number_cell_positions for item in sublist]
join_damagename_cell_positions = [item for sublist in damagename_cell_positions for item in sublist]
join_lank_cell_positions = [item for sublist in lank_cell_positions for item in sublist]
join_picture_cell_positions = [item for sublist in picture_cell_positions for item in sublist]
join_lasttime_lank_cell_positions = [item for sublist in lasttime_lank_cell_positions for item in sublist]
join_damage_memo_cell_positions = [item for sublist in damage_memo_cell_positions for item in sublist]
# セル位置のリストを生成 ↑

number_change = {
'①': '腐食',
'②': '亀裂',
'③': 'ゆるみ・脱落',
'④': '破断',
'⑤': '防食機能の劣化',
'⑥': 'ひびわれ',
'⑦': '剥離・鉄筋露出',
'⑧': '漏水・遊離石灰',
'⑨': '抜け落ち',
'⑩': '補修・補強材の損傷',
'⑪': '床版ひびわれ',
'⑫': 'うき',
'⑬': '遊間の異常',
'⑭': '路面の凹凸',
'⑮': '舗装の異常',
'⑯': '支承部の機能障害',
'⑰': 'その他',
'⑱': '定着部の異常',
'⑲': '変色・劣化',
'⑳': '漏水・滞水',
'㉑': '異常な音・振動',
'㉒': '異常なたわみ',
'㉓': '変形・欠損',
'㉔': '土砂詰まり',
'㉕': '沈下・移動・傾斜',
'㉖': '洗掘',
}

""" 写真の動作 """
# 最大の写真サイズ（幅、高さ）
max_width, max_height = 240, 180 # 4:3

# 位置を追跡するカウンタ
cell_counter = 0

def convert_backslash_to_slash(path):
    return path.replace("/", "\\")

for item in output_data:
    # 3列目(インデックスが2)でpictureキーに写真が設定されている場合
    if cell_counter % 2 == 0 and item['last_time_picture'] is not None:
        # 3列目を空白にするため、インデックスを1つ追加
        cell_counter += 1
    # pictureキーに写真が設定されていても、3列目でなければOK
    else:
        pass
    
    if cell_counter % 6 == 5 and cell_counter > 10:
        # プログラム4｜マクロ実行
        hide_sheet_copy_and_paste(wb, ws)

    # 現在の画像を貼り付ける動作
    for image_path in item['this_time_picture']:
        full = "C:\work\django\myproject\program\Infraproject\infra\static"
        sub_image_path_full = os.path.join(full, image_path) # フルパスを取得
        image_path_full = convert_backslash_to_slash(sub_image_path_full)
        print(image_path_full)
        if os.path.exists(image_path_full):
            print('true')
            print(image_path_full)
            cell_pos = picture_cell_positions[cell_counter // len(picture_columns)][cell_counter % len(picture_columns)]  # 所定のセル位置
            print(f"cell_pos:{cell_pos}")
            # 画像を開いてリサイズ
            pil_img = pil_img = PILImage.open(image_path_full)
            width, height = pil_img.size
            aspect_ratio = width / height

            if aspect_ratio > max_width / max_height:
                new_width = min(width, max_width)
                new_height = new_width / aspect_ratio
            else:
                new_height = min(height, max_height)
                new_width = new_height * aspect_ratio

            resized_img = pil_img.resize((int(new_width), int(new_height)))

            # 一時ファイル用のテンポラリディレクトリを作成
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                resized_img_path = tmp.name

            # 画像を一時ファイルに保存
            resized_img.save(resized_img_path)

            # openpyxl用の画像オブジェクトを作成
            img = OpenpyxlImage(resized_img_path)

            # セルの位置に画像を貼り付け
            img.anchor = cell_pos
            ws.add_image(img)
            cell_counter += 1  # カウンタを進める
        else:
            print('false')

    # 過去の画像を貼り付ける動作
    if item['last_time_picture'] and os.path.exists(item['last_time_picture']):
        img_path = os.path.abspath(item['last_time_picture'])
        print(img_path)
        cell_pos = picture_cell_positions[cell_counter // len(picture_columns)][cell_counter % len(picture_columns)]
        print(f"cell_pos:{cell_pos}")
        pil_img = pil_img = PILImage.open(img_path)
        width, height = pil_img.size
        aspect_ratio = width / height

        if aspect_ratio > max_width / max_height:
            new_width = min(width, max_width)
            new_height = new_width / aspect_ratio
        else:
            new_height = min(height, max_height)
            new_width = new_height * aspect_ratio

        resized_img = pil_img.resize((int(new_width), int(new_height)))

        # 一時ファイル用のテンポラリディレクトリを作成
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
            resized_img_path = tmp.name

        # 画像を一時ファイルに保存
        resized_img.save(resized_img_path)

        # openpyxl用の画像オブジェクトを作成
        img = OpenpyxlImage(resized_img_path)

        # セルの位置に画像を貼り付け
        img.anchor = cell_pos
        ws.add_image(img)
        cell_counter += 1  # カウンタを進める
        
""" 写真の動作 """
        
# データの入力
data_index = 0

for item, part_pos, number_pos, name_pos, lank_pos, memo_pos, picture_pos in zip_longest(output_data, join_partsname_cell_positions, join_number_cell_positions, join_damagename_cell_positions, join_lank_cell_positions, join_damage_memo_cell_positions, join_picture_cell_positions, fillvalue=None):
    
    if (data_index == 2 or data_index % 2 == 3) and item['last_time_picture'] is not None:
        data_index += 1

    part_cell = join_partsname_cell_positions[data_index]
    number_cell = join_number_cell_positions[data_index]
    name_cell = join_damagename_cell_positions[data_index]
    post_lank_cell = join_lank_cell_positions[data_index]
    pre_lank_cell = join_lasttime_lank_cell_positions[data_index]
    memo_cell = join_damage_memo_cell_positions[data_index]
    
    print(part_cell) # I10
    
# メモに入れるための固定コード　↓
    # firstキーの内容を所定の書式に変更
    try:
        first_data = item['parts_name'][0][0] # 排水管 Dp0101
        split_space = first_data.split(" ")
        first_part_data = split_space[0] # 排水管

        match = re.search(r'\d+', split_space[1])
        first_number_data = match.group() if match else '' # 0101

        if part_cell:
            ws[part_cell] = first_part_data # 排水管
        if number_cell:
            ws[number_cell] = first_number_data # 0101

        second_data = item['damage_name'][0][0] # ①腐食(小小)-b
        second_damage_name = second_data[0] # ①
        second_name_data = number_change.get(second_damage_name, second_damage_name) # 腐食
        second_lank_data = second_data[-1] # b

        if name_cell:
            ws[name_cell] = second_name_data # 腐食
        if post_lank_cell:
            ws[post_lank_cell] = second_lank_data # b

        memo_data = item['textarea_content'] # 損傷メモ 
        if memo_cell:
            ws[memo_cell] = memo_data

        if item['last_time_picture'] is not None:
            data_index += 2
        else:
            data_index += 1
            
    except (TypeError, KeyError):
        if part_cell:
            ws[part_cell] = ""
        if number_cell:
            ws[number_cell] = ""
        if name_cell:
            ws[name_cell] = ""
        if post_lank_cell:
            ws[post_lank_cell] = ""
        if memo_cell:
            ws[memo_cell] = ""
        
# メモに入れるための固定コード　↑

# 現在の日時を取得してファイル名に追加
timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
# 新しいファイル名の生成
new_filename = f"{timestamp}_Macro_{original_file_path}"
# デスクトップのパス
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
# 保存するファイルのフルパス
save_path = os.path.join(desktop_path, new_filename)

# 保存する
wb.save(save_path)
print(f"複製シートが保存されました: {save_path}")