import openpyxl
from openpyxl.styles import NamedStyle

# Workbookを作成
wb = openpyxl.load_workbook('work.xlsx')
ws = wb.worksheets[1]

# セルに数値を入力
ws['F5'].value = "フリガナハシ"
ws['F6'].value = "漢字橋"
ws['H7'].value = "所在地自"
ws['H8'].value = {{ object.span_number }}
ws['AG5'].value = 123 # 路線番号
ws['AG2'].value = 35 # 起点側緯度
ws['AJ2'].value = 39
ws['AL2'].value = 52.5
ws['AG3'].value = 140 # 起点側経度
ws['AJ3'].value = 52
ws['AL3'].value = 11.3
ws['AX2'].value = 52.5 # 終点側緯度
ws['AX3'].value = 32.4 # 終点側経度
ws['BC7'].value = "2023/11/11"
ws['H10'].value = 1992
ws['H11'].value = "2径間連続RCT桁"
ws['H13'].value = "逆T式橋台(A1橋台,A2橋台)T型橋脚(P1橋脚)"
ws['H15'].value = "直接基礎(A1橋台,A2橋台)杭基礎(P1橋脚)"
ws['X10'].value = "TL-20"
ws['AD10'].value = "一等橋"
ws['AK10'].value = "昭和53年 道路橋示方書"
ws['H19'].value = '=$O$29'


# 数式セルに数式スタイルを適用
style = NamedStyle(name='formula', number_format='0')

# ファイルを保存
wb.save('work.xlsx')